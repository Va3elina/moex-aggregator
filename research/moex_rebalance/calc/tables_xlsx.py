#!/usr/bin/env python
"""Те же три таблицы в Excel (../TABLES.xlsx) с живыми формулами: входы на листе «Входы», веса новой базы
считаются на листе «Новая база», остальное ссылается на них."""
import json, pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
BASE = pathlib.Path(__file__).resolve().parents[1]; D = BASE/"data"; A = "2026-09-01"
iss = json.loads((D/"iss_imoex.json").read_text(encoding="utf-8"))["бумаги"]
ffn = json.loads((D/"ff_new.json").read_text(encoding="utf-8")); lwn = json.loads((D/"lw_new.json").read_text(encoding="utf-8"))
tw = json.loads((D/"trades_by_weight.json").read_text(encoding="utf-8"))["веса"]
pr = json.loads((D/"prices.json").read_text(encoding="utf-8")); funds = json.loads((D/"funds.json").read_text(encoding="utf-8"))["сча"]
TV = json.loads((D/"turnover_20d.json").read_text(encoding="utf-8"))["средний_дневной_оборот_руб"]
LWC = {"LKOH": 1.0, "SBER": 0.3, "SBERP": 0.6}; EXCL = {"LENT", "MSNG"}
NAMES = {"LKOH":"ЛУКОЙЛ","SBER":"Сбербанк ао","SBERP":"Сбербанк ап","GAZP":"Газпром","TATN":"Татнефть ао","TATNP":"Татнефть ап","LENT":"Лента","MSNG":"Мосэнерго","YDEX":"Яндекс","T":"Т-Технологии","GMKN":"Норникель","NVTK":"НОВАТЭК","ROSN":"Роснефть","X5":"ИКС 5","OZON":"Озон","VTBR":"ВТБ","SNGSP":"Сургутнефтегаз ап","SNGS":"Сургутнефтегаз ао","PLZL":"Полюс","HEAD":"Хэдхантер","MOEX":"МосБиржа","MTSS":"МТС","IRAO":"Интер РАО","RUAL":"Русал","CHMF":"Северсталь","CBOM":"МКБ","NLMK":"НЛМК","PHOR":"ФосАгро","DOMRF":"Дом.РФ","SVCB":"Совкомбанк","MAGN":"ММК","TRNFP":"Транснефть ап","RTKM":"Ростелеком","AFLT":"Аэрофлот","ENPG":"Эн+","MDMG":"Мать и дитя","BSPB":"Банк СПб","FLOT":"Совкомфлот","ALRS":"АЛРОСА","CNRU":"Циан","POSI":"Позитив","UGLD":"ЮГК","AFKS":"АФК Система","RAGR":"Русагро","VKCO":"VK","RENI":"Ренессанс"}
F = "Arial"; hdr_fill = PatternFill("solid", fgColor="DDE3EA"); inp = Font(name=F, color="0000FF"); bold = Font(name=F, bold=True); thin = Side(style="thin", color="BBBBBB")
wb = Workbook()
def sheet(title, headers, widths):
    ws = wb.create_sheet(title); ws.append(headers)
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]: c.font = bold; c.fill = hdr_fill; c.alignment = Alignment(wrap_text=True, vertical="center"); c.border = Border(bottom=thin)
    ws.row_dimensions[1].height = 42; ws.freeze_panes = "C2"; return ws
# ---------- Входы ----------
wi = wb.active; wi.title = "Входы"; wi.column_dimensions["A"].width = 46; wi.column_dimensions["B"].width = 20; wi.column_dimensions["C"].width = 60
rows = [("Дата фиксации весов новой базы", "31.08.2026", "последний торговый день месяца перед вступлением — так подписаны файлы биржи «Новые базы расчета» (Weight (30.05.2025) и т.д.)"),
        ("Дата цен и капитализаций", "01.09.2026", "ISS МосБиржи, состав IMOEX"),
        ("Потолок веса эмитента", 0.15, "Приложение 3 методики"),
        ("СЧА EQMX, ₽", funds["EQMX"]["руб"], "на 31.08.2026"), ("СЧА TMOS, ₽", funds["TMOS"]["руб"], "на 31.08.2026"), ("СЧА SBMX, ₽", funds["SBMX"]["руб"], "на 31.08.2026"),
        ("СЧА трёх фондов, ₽", "=B4+B5+B6", "сумма"),
        ("Σ учитываемой капитализации СЕЙЧАС, ₽", "=SUMPRODUCT('Новая база'!D2:D47,'Новая база'!E2:E47,'Новая база'!G2:G47)", "cap × FF было × W было по всем 46"),
        ("Σ учитываемой капитализации НОВОЙ базы, ₽", "=SUM('Новая база'!L2:L47)", "cap × FF новый × W новый по 44 оставшимся"),
        ("Множитель веса для нетронутых бумаг", "=B8/B9", "= 1 + доля освободившегося веса; так распределяется вес ушедших и срезанных потолком")]
for r in rows: wi.append(r)
for r in wi.iter_rows(min_row=1, max_row=wi.max_row):
    for c in r: c.font = Font(name=F)
for a in ("B2","B3","B4","B5","B6"): wi[a].font = inp
wi["B3"].number_format = "0%"; 
for a in ("B4","B5","B6","B7","B8","B9"): wi[a].number_format = "#,##0"
wi["B10"].number_format = "0.0000"
wi["A12"] = "Синим — исходные числа, чёрным — формулы. Веса новой базы считаются на листе «Новая база», остальные листы ссылаются на них."; wi["A12"].font = Font(name=F, italic=True)
# ---------- Новая база ----------
nb = sheet("Новая база", ["Тикер","Название","Цена 01.09, ₽","Капитализация, ₽","FF было","FF станет","W было (WW×LW)","LW было","LW станет","WW станет","W станет","MC новой базы, ₽","Вес станет","Вес сейчас (ISS)"],
           [8,20,12,18,9,9,13,9,9,12,12,18,11,12])
order = sorted(tw, key=lambda t: -tw[t]["new"]) 
order = [t for t in order if t not in EXCL] + sorted(EXCL)
row_of = {}
for i, t in enumerate(order, 2):
    b = iss[t][A]; row_of[t] = i; gone = t in EXCL
    nb.append([t, NAMES.get(t, t), pr.get(t, {}).get(A), b["cap"], b["ff"], (None if gone else ffn.get(t, b["ff"])), b["w"], LWC.get(t, b["w"]), (None if gone else lwn.get(t, 1.0)), (None if gone else (tw[t].get("ww") or 1.0)),
               (None if gone else f"=J{i}*I{i}"), (0 if gone else f"=D{i}*F{i}*K{i}"), f"=L{i}/Входы!$B$9", b["weight"]/100])
    for c in nb[i]: c.font = Font(name=F)
    for col in ("C","D","E","F","G","H","I","J"): nb[f"{col}{i}"].font = inp
    nb[f"D{i}"].number_format = "#,##0"; nb[f"L{i}"].number_format = "#,##0"; nb[f"C{i}"].number_format = "#,##0.00"
    for col in ("G","J","K"): nb[f"{col}{i}"].number_format = "0.0000000"
    for col in ("E","F","H","I"): nb[f"{col}{i}"].number_format = "0.00"
    for col in ("M","N"): nb[f"{col}{i}"].number_format = "0.000%"
    if gone: nb[f"B{i}"].comment = Comment("Уходит из индекса 18.09.2026", "Frame")
nb["J2"].comment = Comment("WW — коэффициент потолка, у ЛУКОЙЛа и Сбербанка < 1; рассчитан процедурой п. 2.8.4 на 31.08.2026, биржа опубликует свой в файле «Новые базы расчета». У остальных = 1.", "Frame")
nb["H2"].comment = Comment("LW было: у необрезанных = W было (WW=1); у ЛУКОЙЛа 1,0, Сбербанка ао 0,3, ап 0,6 (из истории ISS).", "Frame")
nb["I2"].comment = Comment("LW станет — из PDF-приложения к релизу n103733 (60 бумаг). Кого нет в PDF — 1.", "Frame")
nb["F2"].comment = Comment("FF станет — из таблицы пресс-релиза n103733 от 28.08.2026 (15 бумаг). Остальные без изменений.", "Frame")
last = len(order) + 1
nb.append(["Σ", "", "", f"=SUM(D2:D{last})", "", "", "", "", "", "", "", f"=SUM(L2:L{last})", f"=SUM(M2:M{last})", f"=SUM(N2:N{last})"])
for c in nb[last+1]: c.font = bold
nb[f"M{last+1}"].number_format = "0.000%"; nb[f"N{last+1}"].number_format = "0.000%"; nb[f"L{last+1}"].number_format = "#,##0"; nb[f"D{last+1}"].number_format = "#,##0"
# ---------- 1. Было → станет ----------
s1 = sheet("1 Было-станет", ["Тикер","Название","Вес сейчас","Вес станет","Δ, п.п.","Позиция фондов сейчас, ₽","Позиция станет, ₽","Разница (сделка), ₽","% позиции"], [8,20,11,11,9,20,20,20,10])
order1 = sorted(tw, key=lambda t: tw[t]["mln"])
for i, t in enumerate(order1, 2):
    r = row_of[t]
    s1.append([t, NAMES.get(t, t), f"='Новая база'!N{r}", f"='Новая база'!M{r}", f"=(D{i}-C{i})*100", f"=C{i}*Входы!$B$7", f"=D{i}*Входы!$B$7", f"=G{i}-F{i}", f"=IF(F{i}=0,0,H{i}/F{i})"])
    for c in s1[i]: c.font = Font(name=F)
    for col in ("C","D"): s1[f"{col}{i}"].number_format = "0.000%"
    s1[f"E{i}"].number_format = "+0.000;-0.000"; s1[f"I{i}"].number_format = "+0.0%;-0.0%"
    for col in ("F","G","H"): s1[f"{col}{i}"].number_format = "#,##0"
    s1[f"H{i}"].font = bold
n1 = len(order1) + 1
s1.append(["Продать всего", "", "", "", "", "", "", f"=SUMIF(H2:H{n1},\"<0\")", f"=H{n1+1}/Входы!$B$7"]); s1.append(["Докупить всего", "", "", "", "", "", "", f"=SUMIF(H2:H{n1},\">0\")", f"=H{n1+2}/Входы!$B$7"])
for rr in (n1+1, n1+2):
    for c in s1[rr]: c.font = bold
    s1[f"H{rr}"].number_format = "#,##0"; s1[f"I{rr}"].number_format = "0.00%"
s1[f"I{n1+1}"].comment = Comment("Доля от СЧА трёх фондов", "Frame")
# ---------- 2. Сделка vs free-float ----------
s2 = sheet("2 Сделка vs FF", ["Тикер","Название","Капитализация, ₽","FF сейчас","FF станет","Free-float, ₽ (cap × FF станет)","Сделка, ₽","% от free-float","Цена, ₽","Сделка, штук","Ср. дневной оборот, ₽","Сделка, дней оборота"], [8,20,18,9,9,20,16,12,11,14,18,12])
order2 = sorted(tw, key=lambda t: -abs(tw[t]["mln"]/(iss[t][A]["cap"]*ffn.get(t, iss[t][A]["ff"]))))
for i, t in enumerate(order2, 2):
    r = row_of[t]; r1 = order1.index(t) + 2; gone = t in EXCL
    s2.append([t, NAMES.get(t, t), f"='Новая база'!D{r}", f"='Новая база'!E{r}", (f"='Новая база'!E{r}" if gone else f"='Новая база'!F{r}"), f"=C{i}*E{i}", f"='1 Было-станет'!H{r1}", f"=G{i}/F{i}", f"='Новая база'!C{r}", f"=IF(I{i}>0,G{i}/I{i},\"\")", TV.get(t), f"=IF(K{i}>0,ABS(G{i})/K{i},\"\")"])
    for c in s2[i]: c.font = Font(name=F)
    s2[f"K{i}"].font = inp
    for col in ("C","F","G","K"): s2[f"{col}{i}"].number_format = "#,##0"
    for col in ("D","E"): s2[f"{col}{i}"].number_format = "0.00"
    s2[f"H{i}"].number_format = "+0.000%;-0.000%"; s2[f"H{i}"].font = bold; s2[f"I{i}"].number_format = "#,##0.00"; s2[f"J{i}"].number_format = "+#,##0;-#,##0"; s2[f"L{i}"].number_format = "0.00"
s2[f"A{len(order2)+3}"] = "У Ленты и Мосэнерго free-float по старому FF (нового нет — бумаги уходят). Оборот — средний дневной за 04.08–01.09.2026 по дневным свечам МосБиржи."; s2[f"A{len(order2)+3}"].font = Font(name=F, italic=True)
# ---------- Газпром ----------
g = wb.create_sheet("Газпром по шагам"); g.column_dimensions["A"].width = 6; g.column_dimensions["B"].width = 78; g.column_dimensions["C"].width = 18
rg = row_of["GAZP"]; rl = row_of["LKOH"]; rs = row_of["SBER"]; rsp = row_of["SBERP"]; rle = row_of["LENT"]; rm = row_of["MSNG"]
def w(t): return f"'Новая база'!N{row_of[t]}"
def wn(t): return f"'Новая база'!M{row_of[t]}"
steps = [("Шаг","Что считаем","Число"),
 (1, "Собственные коэффициенты Газпрома не менялись: FF было → станет, LW было → станет. Множитель = (FF станет × W станет) / (FF было × W было)", f"=('Новая база'!F{rg}*'Новая база'!K{rg})/('Новая база'!E{rg}*'Новая база'!G{rg})"),
 (2, "Освободившийся вес, п.п.: Лента + Мосэнерго + (ЛУКОЙЛ сейчас − станет) + (Сбербанк ао+ап сейчас − станет)", f"=({w('LENT')}+{w('MSNG')}+({w('LKOH')}-{wn('LKOH')})+({w('SBER')}-{wn('SBER')})+({w('SBERP')}-{wn('SBERP')}))*100"),
 (3, "Он достаётся всем бумагам, которых потолок не режет, пропорционально весу. Их суммарный вес сейчас, %", f"=(1-{w('LKOH')}-{w('SBER')}-{w('SBERP')}-{w('LENT')}-{w('MSNG')})*100"),
 (4, "Множитель веса для нетронутой бумаги (лист Входы, B10) и прибавка к позиции, %", "=(Входы!B10-1)*100"),
 (5, "Вес Газпрома сейчас, %", f"={w('GAZP')}*100"),
 (6, "Прибавка Газпрому, п.п. = вес сейчас × прибавка", "=C6*C5/100"),
 (7, "× СЧА трёх фондов = покупка, ₽", "=C7/100*Входы!B7"),
 (8, "Контроль: та же покупка из таблицы 1, ₽", f"='1 Было-станет'!H{order1.index('GAZP')+2}"),
 (9, "Для сравнения Яндекс: вес сейчас × та же прибавка × СЧА, ₽", f"={w('YDEX')}*C5/100*Входы!B7"),
 (10, "Вывод", "Газпром получает тот же процент, что и все нетронутые бумаги. Он просто самая большая позиция среди тех, кому раздают освободившийся вес.")]
for s in steps: g.append(list(s))
for r in g.iter_rows(min_row=1, max_row=g.max_row):
    for c in r: c.font = Font(name=F); c.alignment = Alignment(wrap_text=True, vertical="top")
for c in g[1]: c.font = bold; c.fill = hdr_fill
g["C2"].number_format = "0.0000"; g["C3"].number_format = "0.00"; g["C4"].number_format = "0.0"; g["C5"].number_format = "0.00"; g["C6"].number_format = "0.000"; g["C7"].number_format = "0.000"
for a in ("C8","C9","C10"): g[a].number_format = "#,##0"
wb.move_sheet("Входы", offset=-10)
out = BASE/"TABLES.xlsx"; wb.save(out); print("сохранён", out)
