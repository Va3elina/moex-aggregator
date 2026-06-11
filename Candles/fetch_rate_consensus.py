#!/usr/bin/env python3
"""Консенсус-прогноз ключевой ставки (МО Банка России) → macro_data.

ЦБ ежемесячно публикует «Макроэкономический опрос Банка России» (mo_br) —
сводные статистики прогнозов аналитиков. Берём МЕДИАНУ прогноза ключевой ставки:

  - лист '3'  — ключевая ставка (% годовых, в среднем за год) по ГОДОВЫМ
                горизонтам (2021..2029): indicator RATE_CONSENSUS_<year>;
  - лист '16' — нейтральная (долгосрочная / long-run) ключевая ставка:
                indicator RATE_CONSENSUS_LONGRUN.

period_date = дата раунда опроса (1-е число месяца). Получается помесячный
ряд ОЖИДАНИЙ ставки. Существующий KEY_RATE (реализованная ставка) НЕ трогаем.

Источник публичный, без авторизации. ID файла в URL (144490) ротируется при
обновлении опроса — резолвим динамически со страницы mo_br.

Структура листов транспонирована и отличается между листом 3 и 16 — см.
docstring у parse_year_horizons() / parse_neutral_rate().

ВАЖНО (reachability): cbr.ru исторически нестабилен с московского прод-сервера
(cbr_orfr поэтому на ручном ингесте). На момент написания mo_br ОТВЕЧАЕТ с
сервера (~0.1с). Если в момент ингеста таймаутит — гнать парс с внешней машины
и слать строки оркестратору (паттерн cbr_orfr).

Запуск (ингест в БД): python Candles/fetch_rate_consensus.py
Dry-run без записи:    python Candles/fetch_rate_consensus.py --dry-run
"""
import io
import os
import re
import sys
import argparse
import urllib.request
from collections import Counter
from datetime import date

import openpyxl
from sqlalchemy import create_engine, text

DB_URL = os.getenv("DB_URL")
UA = "Mozilla/5.0 (compatible; FrameBot/1.0)"
PAGE = "https://www.cbr.ru/statistics/ddkp/mo_br/"
XLSX_FALLBACK = "https://www.cbr.ru/Content/Document/File/144490/full.xlsx"
SOURCE = "CBR mo_br"


def fetch_bytes(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=60) as r:
        return r.read()


def resolve_xlsx() -> str:
    """ID файла (144490) ротируется при обновлении опроса — резолвим со страницы."""
    html = fetch_bytes(PAGE).decode("utf-8", "ignore")
    m = re.findall(r'href="(/Content/Document/File/\d+/full\.xlsx)"', html)
    return "https://www.cbr.ru" + m[0] if m else XLSX_FALLBACK


def _round_columns(header_row, start_col):
    """[(col_index, date), ...] из строки-шапки с датами раундов (datetime в ячейках)."""
    out = []
    for ci in range(start_col, len(header_row)):
        v = header_row[ci]
        if hasattr(v, "year"):
            out.append((ci, date(v.year, v.month, v.day)))
    return out


def parse_year_horizons(wb):
    """Лист '3': медиана средней ставки за год по годовым горизонтам × раундам.

    Структура транспонирована: строки = блоки статистик (метка в кол. B, idx 1,
    forward-fill вниз; англ. в C, idx 2), внутри блока строка на год-горизонт
    (дата конца года в кол. D, idx 3); столбцы раундов — с кол. E (idx 4),
    даты раундов в строке 6 (rows[5]). Пропуски = '-'.
    """
    ws = wb["3"]
    rows = list(ws.iter_rows(values_only=True))
    round_cols = _round_columns(rows[5], start_col=4)

    recs = []
    cur_stat = None
    for r in rows:
        label = r[1].strip() if isinstance(r[1], str) else None
        if label:
            cur_stat = label  # forward-fill метки статистики вниз по блоку
        if cur_stat != "Медиана":
            continue
        yhz = r[3]
        if not hasattr(yhz, "year"):
            continue
        horizon_year = yhz.year
        for ci, rdate in round_cols:
            val = r[ci] if ci < len(r) else None
            if val in (None, "-", ""):
                continue
            try:
                fval = float(val)
            except (TypeError, ValueError):
                continue
            recs.append({
                "indicator": f"RATE_CONSENSUS_{horizon_year}",
                "name": (f"Консенсус-прогноз ключевой ставки на {horizon_year} "
                         f"(медиана, средняя за год; МО Банка России)"),
                "period_date": rdate,
                "value": round(fval, 2),
            })
    return recs


def parse_neutral_rate(wb):
    """Лист '16': нейтральная (долгосрочная / long-run) ключевая ставка.

    Структура ОТЛИЧАЕТСЯ от листа '3': плоский одиночный ряд. Даты раундов —
    в строке 6 (rows[5]) с кол. D (idx 3); строка медианы помечена 'Медиана' в
    кол. B, значения с той же кол. D (idx 3). Год-горизонта нет — горизонт
    «долгосрочный». Пропуски = '-'.
    """
    ws = wb["16"]
    rows = list(ws.iter_rows(values_only=True))
    round_cols = _round_columns(rows[5], start_col=3)

    median_row = None
    for r in rows:
        if isinstance(r[1], str) and r[1].strip() == "Медиана":
            median_row = r
            break
    if median_row is None:
        return []

    recs = []
    for ci, rdate in round_cols:
        val = median_row[ci] if ci < len(median_row) else None
        if val in (None, "-", ""):
            continue
        try:
            fval = float(val)
        except (TypeError, ValueError):
            continue
        recs.append({
            "indicator": "RATE_CONSENSUS_LONGRUN",
            "name": ("Консенсус-прогноз нейтральной (долгосрочной) ключевой ставки "
                     "(медиана; МО Банка России)"),
            "period_date": rdate,
            "value": round(fval, 2),
        })
    return recs


def collect():
    url = resolve_xlsx()
    print(f"resolved xlsx: {url}")
    data = fetch_bytes(url)
    print(f"xlsx bytes: {len(data)}")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    recs = parse_year_horizons(wb) + parse_neutral_rate(wb)
    if not recs:
        print("FATAL: rate_consensus не распознан", file=sys.stderr)
        sys.exit(1)
    return recs


def summarize(recs):
    c = Counter(x["indicator"] for x in recs)
    print(f"rate_consensus: {len(recs)} строк, {len(c)} индикаторов")
    for ind, n in sorted(c.items()):
        print(f"  {ind}: {n}")


def upsert(recs):
    if not DB_URL:
        print("FATAL: DB_URL не задан в env", file=sys.stderr)
        sys.exit(1)
    eng = create_engine(DB_URL)
    # имя индикатора → берём самое свежее name (горизонтные одинаковы внутри indicator)
    names = {x["indicator"]: x["name"] for x in recs}
    starts = {}
    for x in recs:
        starts[x["indicator"]] = min(starts.get(x["indicator"], x["period_date"]),
                                     x["period_date"])

    with eng.begin() as conn:
        for ind, nm in names.items():
            conn.execute(text("""
                INSERT INTO macro (indicator, name, frequency, source, start_date)
                VALUES (:i, :n, 'monthly', :src, :sd)
                ON CONFLICT (indicator) DO UPDATE
                  SET name = EXCLUDED.name,
                      start_date = LEAST(macro.start_date, EXCLUDED.start_date)
            """), {"i": ind, "n": nm, "src": SOURCE, "sd": starts[ind]})

        n = 0
        for x in recs:
            conn.execute(text("""
                INSERT INTO macro_data (indicator, period_date, value, source)
                VALUES (:i, :d, :v, :src)
                ON CONFLICT (indicator, period_date) DO UPDATE
                  SET value = EXCLUDED.value
            """), {"i": x["indicator"], "d": x["period_date"],
                   "v": x["value"], "src": SOURCE})
            n += 1
    print(f"upsert OK: {n} строк macro_data")


def main():
    ap = argparse.ArgumentParser(description="Консенсус-прогноз ключевой ставки (МО ЦБ) → macro_data")
    ap.add_argument("--dry-run", action="store_true",
                    help="только fetch+parse+print, без записи в БД")
    args = ap.parse_args()

    recs = collect()
    summarize(recs)

    if args.dry_run:
        print("\n-- DRY-RUN: первые 8 строк --")
        for x in recs[:8]:
            print(f"   {x['indicator']}  {x['period_date']}  {x['value']}")
        return

    upsert(recs)


if __name__ == "__main__":
    main()
