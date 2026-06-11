#!/usr/bin/env python3
"""Результаты аукционов ОФЗ Минфина РФ -> таблица ofz_auctions.

Минфин публикует годовой сводный XLSX с результатами первичных размещений ОФЗ
(аукционы + доразмещения ДРПА). Файл текущего года обновляется еженедельно по
средам; на странице-листинге висят файлы за 2021..текущий год (полная история
по году на файл). Cbonds для аукционов НЕ подходит (mobile-API не отдаёт датасет,
web-API 401), MOEX ISS борд AUCT пуст по бумагам - Минфин XLSX единственный
полноценный источник.

Источник (листинг, откуда дискаверим ссылку на свежий файл):
  https://minfin.gov.ru/ru/perfomance/public_debt/internal/operations/ofz/auction/
Имя файла плавает по as-of дате:
  INTERNET_Auction_Results_rus_<ГОД>_<asof>.xlsx
-> скрейпим href по regex, берём для каждого года максимальную as-of дату.

ДРЕЙФ ФОРМАТА (важно):
  * 2021-2023: 14 колонок, БЕЗ колонки Формат (все строки = обычный Аукцион).
  * 2024+:     15 колонок, добавлена колонка 2 Формат* (Аукцион | ДРПА).
  Лэйаут определяется автоматически по позиции колонки Код выпуска (NNNNNRMFS):
  индекс 2 -> 15-колоночный файл с Формат, индекс 1 -> 14-колоночный.

GOTCHA по ключу: в 15-колоночных файлах один и тот же (auction_date, secid)
встречается ДВАЖДЫ - обычный Аукцион и в тот же день ДРПА (доразмещение).
Поэтому PK включает auction_type: (auction_date, secid, auction_type).

Все руб-объёмы в источнике - в МЛН рублей; храним как есть в *_mln (точность).
Пустые/плейсхолдер значения ('-', '-***', '–', '—') -> NULL (в т.ч. demand у ДРПА).

Запуск (реальный ингест - делает оркестратор-человек):
  python Candles/fetch_ofz_auctions.py                # все годы с листинга
  python Candles/fetch_ofz_auctions.py --year 2026    # только один год
  python Candles/fetch_ofz_auctions.py --dry-run      # парс+печать, БЕЗ записи в БД
"""
import os
import io
import re
import sys
import argparse
import urllib.request
from datetime import datetime

import openpyxl

DB_URL = os.getenv("DB_URL")
UA = "Mozilla/5.0 (compatible; FrameBot/1.0)"
LISTING = ("https://minfin.gov.ru/ru/perfomance/public_debt/"
           "internal/operations/ofz/auction/")
HOST = "https://minfin.gov.ru"
FILE_RE = re.compile(r'INTERNET_Auction_Results_rus_(\d{4})_(\d{8})\.xlsx')
HREF_RE = re.compile(r'href="([^"]*INTERNET_Auction_Results_rus_\d{4}_\d{8}\.xlsx)"')
REGNUM_RE = re.compile(r'^\d{5}RMFS$', re.I)
SOURCE = "minfin"


def http_get(url, timeout=60):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def num(x):
    """'-'/'-***'/None/'' -> None; иначе float (с нормализацией запятой/пробелов)."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s in ("", "-", "–", "—") or s.startswith("-*"):
        return None
    try:
        return float(s.replace(",", ".").replace("\xa0", "").replace(" ", ""))
    except ValueError:
        return None


def discover_files(only_year=None):
    """Скрейп листинга -> {год: url} с максимальной as-of датой на год."""
    html = http_get(LISTING).decode("utf-8", "ignore")
    best = {}   # year -> (asof, url)
    for href in HREF_RE.findall(html):
        m = FILE_RE.search(href)
        if not m:
            continue
        year, asof = m.group(1), m.group(2)
        if only_year and year != str(only_year):
            continue
        url = href if href.startswith("http") else HOST + href
        if year not in best or asof > best[year][0]:
            best[year] = (asof, url)
    return {y: v[1] for y, v in sorted(best.items())}


def parse_xlsx(raw):
    """Парс одного годового XLSX -> list[dict]. Авто-детект 14/15-колоночного лэйаута."""
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(values_only=True) if r]

    code_col = None
    for ci in (2, 1):
        if any(len(r) > ci and isinstance(r[ci], str) and REGNUM_RE.match(r[ci].strip())
               for r in rows):
            code_col = ci
            break
    if code_col is None:
        return []
    has_fmt = (code_col == 2)

    out = []
    for r in rows:
        if len(r) <= code_col:
            continue
        d, code = r[0], r[code_col]
        if not isinstance(d, datetime):
            continue
        if not (isinstance(code, str) and REGNUM_RE.match(code.strip())):
            continue

        if has_fmt:
            fmt = str(r[1]).strip() if r[1] else "Аукцион"
            sectype, matdate, days = r[3], r[4], r[5]
            planned, cut_price, avg_price = r[6], r[7], r[8]
            cut_yield, avg_yield = r[9], r[10]
            demand, placed, proceeds, kcoef = r[11], r[12], r[13], (r[14] if len(r) > 14 else None)
        else:
            fmt = "Аукцион"
            sectype, matdate, days = r[2], r[3], r[4]
            planned, cut_price, avg_price = r[5], r[6], r[7]
            cut_yield, avg_yield = r[8], r[9]
            demand, placed, proceeds, kcoef = r[10], r[11], r[12], (r[13] if len(r) > 13 else None)

        regnum = code.strip().upper()
        days_v = num(days)
        out.append({
            "auction_date": d.date(),
            "secid": "SU" + regnum,
            "regnumber": regnum,
            "auction_type": fmt,
            "sectype": str(sectype).strip() if sectype else None,
            "mat_date": matdate.date() if isinstance(matdate, datetime) else None,
            "days_to_mat": int(days_v) if days_v is not None else None,
            "planned_mln": num(planned),
            "placed_mln": num(placed),
            "demand_mln": num(demand),
            "proceeds_mln": num(proceeds),
            "cut_price": num(cut_price),
            "avg_price": num(avg_price),
            "cut_yield": num(cut_yield),
            "avg_yield": num(avg_yield),
            "bid_to_cover": num(kcoef),
        })
    return out


DDL = """
CREATE TABLE IF NOT EXISTS ofz_auctions (
    auction_date  date         NOT NULL,
    secid         varchar(20)  NOT NULL,
    regnumber     varchar(12),
    auction_type  varchar(16)  NOT NULL DEFAULT 'Аукцион',
    issue_name    varchar(64),
    sectype       varchar(16),
    mat_date      date,
    days_to_mat   int,
    planned_mln   numeric(18,3),
    placed_mln    numeric(18,3),
    demand_mln    numeric(18,3),
    proceeds_mln  numeric(18,3),
    cut_price     numeric(10,4),
    avg_price     numeric(10,4),
    cut_yield     numeric(8,4),
    avg_yield     numeric(8,4),
    bid_to_cover  numeric(10,6),
    source        varchar(32)  DEFAULT 'minfin',
    updated_at    timestamptz  DEFAULT now(),
    PRIMARY KEY (auction_date, secid, auction_type)
)
"""

UPSERT = """
INSERT INTO ofz_auctions (
    auction_date, secid, regnumber, auction_type, sectype, mat_date, days_to_mat,
    planned_mln, placed_mln, demand_mln, proceeds_mln,
    cut_price, avg_price, cut_yield, avg_yield, bid_to_cover, source, updated_at
) VALUES (
    :auction_date, :secid, :regnumber, :auction_type, :sectype, :mat_date, :days_to_mat,
    :planned_mln, :placed_mln, :demand_mln, :proceeds_mln,
    :cut_price, :avg_price, :cut_yield, :avg_yield, :bid_to_cover, :source, now()
)
ON CONFLICT (auction_date, secid, auction_type) DO UPDATE SET
    regnumber    = EXCLUDED.regnumber,
    sectype      = EXCLUDED.sectype,
    mat_date     = EXCLUDED.mat_date,
    days_to_mat  = EXCLUDED.days_to_mat,
    planned_mln  = EXCLUDED.planned_mln,
    placed_mln   = EXCLUDED.placed_mln,
    demand_mln   = EXCLUDED.demand_mln,
    proceeds_mln = EXCLUDED.proceeds_mln,
    cut_price    = EXCLUDED.cut_price,
    avg_price    = EXCLUDED.avg_price,
    cut_yield    = EXCLUDED.cut_yield,
    avg_yield    = EXCLUDED.avg_yield,
    bid_to_cover = EXCLUDED.bid_to_cover,
    source       = EXCLUDED.source,
    updated_at   = now()
"""


def fetch_all(only_year=None):
    files = discover_files(only_year)
    if not files:
        print("FATAL: на листинге Минфина не найдено INTERNET_Auction_Results_rus_*.xlsx",
              file=sys.stderr)
        sys.exit(1)
    rows, used = [], []
    for year, url in files.items():
        raw = http_get(url)
        part = parse_xlsx(raw)
        used.append((year, url, len(part)))
        rows.extend(part)
    return rows, used


def print_preview(rows, used):
    print("Использованные годовые файлы Минфина:")
    for year, url, n in used:
        print(f"  {year}: {n:3d} строк  {url}")
    if not rows:
        print("Распарсено 0 строк.", file=sys.stderr)
        return
    aukts = sum(1 for x in rows if x["auction_type"] == "Аукцион")
    print(f"\nВсего распарсено: {len(rows)} строк (Аукцион={aukts}, ДРПА={len(rows) - aukts})")
    print(f"Диапазон дат: {min(x['auction_date'] for x in rows)} .. "
          f"{max(x['auction_date'] for x in rows)}")
    print("\nПервые 8 строк:")
    for x in rows[:8]:
        print(f"  {x['auction_date']} {x['auction_type']:8s} {x['secid']:12s} "
              f"{(x['sectype'] or '?'):7s} plan={x['planned_mln']} "
              f"placed={x['placed_mln']} demand={x['demand_mln']} "
              f"avg_yield={x['avg_yield']} avg_price={x['avg_price']} "
              f"k={x['bid_to_cover']}")


def main():
    ap = argparse.ArgumentParser(description="Ингест результатов аукционов ОФЗ Минфина")
    ap.add_argument("--year", type=int, default=None,
                    help="только один год (по умолчанию - все годы с листинга)")
    ap.add_argument("--dry-run", action="store_true",
                    help="только fetch+parse+печать, БЕЗ записи в БД")
    args = ap.parse_args()

    rows, used = fetch_all(args.year)
    print_preview(rows, used)

    if args.dry_run:
        print("\n[DRY-RUN] В БД ничего не записано.")
        return
    if not rows:
        print("Нечего писать в БД.", file=sys.stderr)
        sys.exit(1)
    if not DB_URL:
        print("FATAL: DB_URL не задан в окружении.", file=sys.stderr)
        sys.exit(1)

    from sqlalchemy import create_engine, text
    eng = create_engine(DB_URL)
    with eng.begin() as conn:
        conn.execute(text(DDL))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_ofz_auctions_secid "
                          "ON ofz_auctions (secid)"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS idx_ofz_auctions_date "
                          "ON ofz_auctions (auction_date)"))
        n = 0
        for rec in rows:
            rec["source"] = SOURCE
            conn.execute(text(UPSERT), rec)
            n += 1
    print(f"\nofz_auctions: upsert {n} строк (источник={SOURCE}).")


if __name__ == "__main__":
    main()