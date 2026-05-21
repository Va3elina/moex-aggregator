#!/usr/bin/env python3
"""
Backfill всех исторических ORFR XLSX с landing page Банка России.

Скачивает все 32 XLSX (от charts_ORFR_2023-04 до ORFR_2026-4) и парсит.

Структура XLSX менялась с годами:

  • 2023-04 → ~2025-08 (старый формат, monthly only):
    - Лист «рис10» / «рис30» / похожий: Акции
    - Лист «рис2»:                       Валюты
    - Лист про ОФЗ может отсутствовать или иметь другой номер
    - Header колонок: «Дата | <категория1> | <категория2> | ...»
    - 8 категорий (есть «Дочерние иностранные организации»)
    - Period — только месяцы (2022-01-01, 2022-02-01, ...)

  • 2025-09 → present (новый формат, quarterly + last month):
    - Лист «рис. 32»: Акции, 7 категорий
    - Лист «рис. 14»: ОФЗ, 7 категорий
    - Лист «рис. 8»:  Валюты, 5 категорий
    - Header: «Годы | Месяцы/кварталы | <категория1> | ...»
    - Period — квартал или последний месяц

Backfill detect format по header и parse соответственно. Через ON CONFLICT
DO UPDATE — самые новые данные перезаписывают дубликаты.

Запуск:
  python3 -m CBR.backfill_orfr_history           # download + parse + upsert all
  python3 -m CBR.backfill_orfr_history --dry-run # parse only, no DB writes
  python3 -m CBR.backfill_orfr_history --year 2023  # filter by year
"""

import argparse
import io
import logging
import os
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
import requests
from sqlalchemy import create_engine, text

sys.path.insert(0, str(Path(__file__).parent.parent))
from dotenv import load_dotenv  # noqa: E402

load_dotenv(Path(__file__).parent.parent / ".env")

DB_URL = os.getenv("DB_URL")
LANDING_URL = "https://cbr.ru/analytics/finstab/orfr/"
BASE_URL = "https://cbr.ru"

# Категории которые ЦБ убрал из методологии — не вставляем в БД.
# КРИТИЧНО: per-instrument! Имена «Нерезиденты», «СЗКО», «Прочие Банки»,
# «Доверительное управление», «Нефинансовые организации» есть И в stocks,
# И в fx. Глобальный set исключал бы их и для акций (баг).
EXCLUDED_BY_TYPE: dict[str, set[str]] = {
    "stocks": {"Дочерние иностранные организации"},
    "ofz": {"Дочерние иностранные организации", "Минфин", "Минфин России"},
    "fx": {
        "Банк России (ЦБ РФ)",
        "Доверительное управление",
        "Нерезиденты",
        "Нефинансовые организации",
        "Прочие Банки",
        "СЗКО",
    },
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────────────────
#  1) Поиск всех XLSX URL'ов
# ──────────────────────────────────────────────────────────────────────────────

XLSX_URL_RE = re.compile(
    r"/Collection/Collection/File/(\d+)/((?:charts_)?ORFR_(\d{4})-?(\d{1,2})?(?:-(\d+))?(?:_(\d{4})-(\d{1,2}))?\.xlsx)",
    re.IGNORECASE,
)


def find_all_xlsx_urls() -> list[dict]:
    """Возвращает список всех XLSX с landing page. Каждый item:
    {url, filename, year, month, file_id}.
    Сортирует chronologically (от старых к новым).
    """
    log.info(f"Получаю landing page {LANDING_URL}…")
    resp = requests.get(LANDING_URL, timeout=30, headers={
        "User-Agent": "Mozilla/5.0 Frame/1.0"
    })
    resp.raise_for_status()
    html = resp.text

    items = []
    seen_ids = set()
    for m in XLSX_URL_RE.finditer(html):
        file_id = m.group(1)
        if file_id in seen_ids:
            continue
        seen_ids.add(file_id)
        filename = m.group(2)
        # Парсим YYYY-M из filename (учитываем «ORFR_2024-25-1» = Dec 2024 / Jan 2025 — берём начальную дату)
        # Самый ранний период = main год+месяц из URL
        # Pattern: charts_ORFR_2024-25-1 → year=2024 month=??
        # Pattern: ORFR_2026-4 → year=2026 month=4
        # Pattern: ORFR_2025-12_2026-01 → year=2025 month=12 (start)
        # Для backfill order — берём конечный месяц как «дата выпуска»
        end_year = int(m.group(6)) if m.group(6) else int(m.group(3))
        end_month_str = m.group(7) if m.group(7) else m.group(4)
        try:
            end_month = int(end_month_str) if end_month_str else 1
        except (TypeError, ValueError):
            end_month = 1
        items.append({
            "file_id": file_id,
            "filename": filename,
            "url": f"{BASE_URL}/Collection/Collection/File/{file_id}/{filename}",
            "year": end_year,
            "month": end_month,
        })

    # FORWARD chronological order (старые → новые) — новые выпуски через
    # ON CONFLICT DO UPDATE перезаписывают старые. ЦБ ревизирует данные
    # прошлых месяцев в свежих выпусках → forward-order даёт АКТУАЛЬНЫЕ
    # (ревизированные) значения. Раньше был reverse для «monthly побеждает
    # quarterly», но кварталы больше не парсятся → reverse не нужен.
    items.sort(key=lambda x: (x["year"], x["month"]))
    log.info(f"  Найдено {len(items)} XLSX")
    return items


# ──────────────────────────────────────────────────────────────────────────────
#  2) Detect format и parse
# ──────────────────────────────────────────────────────────────────────────────

MONTH_RU = {
    "Январь": 1, "Февраль": 2, "Март": 3, "Апрель": 4, "Май": 5, "Июнь": 6,
    "Июль": 7, "Август": 8, "Сентябрь": 9, "Октябрь": 10, "Ноябрь": 11, "Декабрь": 12,
}
QUARTER_RU = {
    "I квартал": 1, "II квартал": 2, "III квартал": 3, "IV квартал": 4,
}


def _quarter_end(year: int, q: int) -> date:
    return date(year, q * 3, [31, 30, 30, 31][q - 1])


def _month_end(year: int, month: int) -> date:
    if month == 12:
        return date(year, 12, 31)
    return date(year, month + 1, 1) - timedelta(days=1)


def _detect_instrument(sheet_title: str) -> Optional[str]:
    """По title листа определяем instrument_type или None.
    Returns: 'stocks' | 'ofz' | 'fx' | None.

    Строгий filter — нужны ОБЕ маркеры:
      1) «по категориям участников» / «по категориям» (показывает что данные
         агрегированы по нашим категориям, не daily/weekly tick data)
      2) ключевое слово инструмента (акции / ОФЗ / валют)

    Это исключает daily charts которые иногда have «нетто» в названии
    но содержат tick data для всех дней месяца.
    """
    t = sheet_title.lower()
    if "по категориям" not in t:
        return None
    # Buy/sell context — должен присутствовать
    has_buy_sell = ("покуп" in t and "продаж" in t) or ("покупатели" in t and "продавц" in t) or ("нетто" in t)
    if not has_buy_sell:
        return None
    if "акци" in t:
        return "stocks"
    if "офз" in t:
        return "ofz"
    if "валют" in t or "иностранн" in t:
        return "fx"
    return None


def find_target_sheets(wb) -> dict[str, str]:
    """Возвращает {instrument_type: sheet_name} для текущего workbook."""
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Title — первая string-ячейка
        title = ""
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and len(cell.strip()) > 5:
                    title = cell
                    break
            if title:
                break
        itype = _detect_instrument(title)
        if itype and itype not in result:
            result[itype] = sheet_name
    return result


def parse_sheet(wb, sheet_name: str, source_file: str, instrument_type: str) -> Iterable[dict]:
    """Yields {period_year, period_label, period_kind, period_end_date, category, value}.

    Robust auto-detect — header может не иметь «Дата»/«Год» label,
    категории определяются как string cells в header row, date column —
    по первому datetime в data row.
    """
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # 1. Find header row — первая row после title rows которая имеет ≥3 строковых ячеек
    header_idx = None
    for idx, r in enumerate(rows[:10]):
        if not r:
            continue
        string_cells = [c for c in r if isinstance(c, str) and c.strip()]
        if len(string_cells) >= 3:
            header_idx = idx
            break
    if header_idx is None or header_idx >= len(rows) - 1:
        log.warning(f"  {sheet_name}: header не найден")
        return

    header = rows[header_idx]

    # 2. Find first data row после header
    data_start = None
    for idx in range(header_idx + 1, min(header_idx + 5, len(rows))):
        r = rows[idx]
        if r and any(c is not None and c != "" for c in r):
            data_start = idx
            break
    if data_start is None:
        log.warning(f"  {sheet_name}: data rows не найдены")
        return

    data_row = rows[data_start]

    # 3. Detect format:
    #    - NEW: header[0]="Год" / "Годы", header[1]="Месяцы/кварталы"
    #    - OLD: первая datetime cell в data row → date column. Категории = string cells в header.
    h0 = header[0] if len(header) > 0 and isinstance(header[0], str) else ""
    h1 = header[1] if len(header) > 1 and isinstance(header[1], str) else ""
    is_new_format = h0.strip() in ("Год", "Годы") and h1.strip() in ("Месяцы/кварталы",) or (
        h0.strip() in ("Год", "Годы") and isinstance(header[1], str) and "квартал" in (h1.lower() if h1 else "")
    )

    if is_new_format:
        category_cols = [(i, c.strip()) for i, c in enumerate(header[2:], start=2)
                         if isinstance(c, str) and c.strip()]
        log.info(f"  {sheet_name}: NEW format, {len(category_cols)} categories")
        yield from _parse_new_format(rows, header_idx, category_cols, source_file, instrument_type)
    else:
        # OLD format. Find date_col by scanning data_row для первого datetime.
        date_col = None
        for i, c in enumerate(data_row):
            if isinstance(c, (date, datetime)):
                date_col = i
                break
        if date_col is None:
            log.warning(f"  {sheet_name}: date column не найден в data row")
            return

        # Категории — все string cells в header AFTER date_col
        category_cols = [(i, c.strip()) for i, c in enumerate(header)
                         if i > date_col and isinstance(c, str) and c.strip()]
        if not category_cols:
            log.warning(f"  {sheet_name}: категории не найдены")
            return
        log.info(f"  {sheet_name}: OLD format (monthly), {len(category_cols)} categories, date_col={date_col}")
        yield from _parse_old_format(rows, header_idx, date_col, category_cols, source_file, instrument_type)


def _parse_new_format(rows, header_idx, category_cols, source_file, instrument_type):
    current_year = None
    count = 0
    for r in rows[header_idx + 1:]:
        if not r or all(c is None or c == "" for c in r[:2]):
            continue
        year_cell, period_cell = r[0], r[1]
        if year_cell is not None and year_cell != "":
            try:
                current_year = int(year_cell)
            except (ValueError, TypeError):
                continue
        if current_year is None or not isinstance(period_cell, str):
            continue
        plabel = period_cell.strip()
        # КВАРТАЛЫ ПРОПУСКАЕМ — индикатор работает только с месячными данными.
        # Квартальная строка имеет тот же period_end_date что последний месяц
        # квартала → затирала бы корректную месячную запись через ON CONFLICT.
        if plabel in QUARTER_RU:
            continue
        elif plabel in MONTH_RU:
            kind = "month"
            end_dt = _month_end(current_year, MONTH_RU[plabel])
        else:
            continue
        excluded = EXCLUDED_BY_TYPE.get(instrument_type, set())
        for col_idx, cat in category_cols:
            if cat in excluded:
                continue
            val = r[col_idx] if col_idx < len(r) else None
            if val is None:
                continue
            try:
                fval = float(val)
            except (ValueError, TypeError):
                continue
            yield {
                "period_year": current_year,
                "period_label": plabel,
                "period_kind": kind,
                "period_end_date": end_dt,
                "category": cat,
                "value": fval,
                "source_file": source_file,
            }
            count += 1


def _parse_old_format(rows, header_idx, date_col, category_cols, source_file, instrument_type):
    """Old format — каждая row имеет date в одной колонке + values в категориях.
    Каждая дата = первый день месяца (2022-01-01), period_end = последний день месяца."""
    count = 0
    for r in rows[header_idx + 1:]:
        if not r or len(r) <= date_col:
            continue
        date_val = r[date_col]
        if date_val is None or date_val == "":
            continue
        # Парсим дату — может быть datetime object или string
        if isinstance(date_val, datetime):
            month_start = date_val.date()
        elif isinstance(date_val, date):
            month_start = date_val
        elif isinstance(date_val, str):
            try:
                month_start = datetime.fromisoformat(date_val).date()
            except (ValueError, TypeError):
                continue
        else:
            continue

        year = month_start.year
        month = month_start.month
        end_dt = _month_end(year, month)
        plabel = list(MONTH_RU.keys())[month - 1]  # «Январь» для month=1

        excluded = EXCLUDED_BY_TYPE.get(instrument_type, set())
        for col_idx, cat in category_cols:
            if cat in excluded:
                continue
            val = r[col_idx] if col_idx < len(r) else None
            if val is None:
                continue
            try:
                fval = float(val)
            except (ValueError, TypeError):
                continue
            yield {
                "period_year": year,
                "period_label": plabel,
                "period_kind": "month",
                "period_end_date": end_dt,
                "category": cat,
                "value": fval,
                "source_file": source_file,
            }
            count += 1


# ──────────────────────────────────────────────────────────────────────────────
#  3) Upsert
# ──────────────────────────────────────────────────────────────────────────────

UPSERT_SQL = text("""
    INSERT INTO cbr_flows (
        instrument_type, period_year, period_label, period_kind,
        period_end_date, category, value, source_file, updated_at
    ) VALUES (
        :instrument_type, :period_year, :period_label, :period_kind,
        :period_end_date, :category, :value, :source_file, NOW()
    )
    ON CONFLICT (instrument_type, period_end_date, category)
    DO UPDATE SET
        value = EXCLUDED.value,
        period_label = EXCLUDED.period_label,
        period_kind = EXCLUDED.period_kind,
        period_year = EXCLUDED.period_year,
        source_file = EXCLUDED.source_file,
        updated_at = NOW()
""")


def upsert_rows(engine, instrument_type: str, rows: list[dict]) -> int:
    if not rows:
        return 0
    # FX в XLSX ЦБ использует обратную конвенцию знаков (см. fetch_orfr_flows.py
    # подробный комментарий). Инвертируем чтобы БД была консистентна:
    # + = покупка, − = продажа для всех instrument_types.
    if instrument_type == "fx":
        rows = [{**r, "value": -r["value"]} for r in rows]
    payload = [{**r, "instrument_type": instrument_type} for r in rows]
    with engine.begin() as conn:
        conn.execute(UPSERT_SQL, payload)
    return len(payload)


# ──────────────────────────────────────────────────────────────────────────────
#  4) Main pipeline
# ──────────────────────────────────────────────────────────────────────────────

def process_xlsx(xlsx_url: str, source_file: str, engine, dry_run: bool) -> dict:
    """Скачивает и парсит один XLSX. Возвращает stats."""
    try:
        resp = requests.get(xlsx_url, timeout=60, headers={"User-Agent": "Mozilla/5.0 Frame/1.0"})
        resp.raise_for_status()
    except Exception as e:
        log.error(f"  Не удалось скачать {xlsx_url}: {e}")
        return {"error": str(e)}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=False)
    except Exception as e:
        log.error(f"  Не удалось открыть XLSX: {e}")
        return {"error": str(e)}

    sheets_map = find_target_sheets(wb)
    if not sheets_map:
        log.warning(f"  Не найдено sheets с данными покупок/продаж")
        return {"sheets": 0}

    stats = {}
    for instrument_type, sheet_name in sheets_map.items():
        rows = list(parse_sheet(wb, sheet_name, source_file, instrument_type))
        if not rows:
            continue
        if dry_run:
            stats[instrument_type] = len(rows)
            log.info(f"    [DRY] {instrument_type}: {len(rows)} rows (range {min(r['period_end_date'] for r in rows)} → {max(r['period_end_date'] for r in rows)})")
        else:
            n = upsert_rows(engine, instrument_type, rows)
            stats[instrument_type] = n
            log.info(f"    ✓ {instrument_type}: upsert {n} rows")
    return stats


def main():
    p = argparse.ArgumentParser(description="Backfill ORFR ЦБ — все исторические XLSX")
    p.add_argument("--dry-run", action="store_true", help="Не писать в БД")
    p.add_argument("--year", type=int, help="Filter: только XLSX за указанный год")
    p.add_argument("--limit", type=int, help="Limit количества XLSX (debug)")
    args = p.parse_args()

    items = find_all_xlsx_urls()
    if args.year:
        items = [it for it in items if it["year"] == args.year]
        log.info(f"  Filter year={args.year}: {len(items)} files")
    if args.limit:
        items = items[: args.limit]
        log.info(f"  Limited to {len(items)} files")

    engine = None if args.dry_run else create_engine(DB_URL, connect_args={"ssl_context": False})

    total = {"stocks": 0, "ofz": 0, "fx": 0, "errors": 0, "files": 0}
    for it in items:
        log.info(f"[{it['year']}-{it['month']:02d}] {it['filename']}")
        source_file = it["filename"].rsplit(".", 1)[0]
        result = process_xlsx(it["url"], source_file, engine, args.dry_run)
        if "error" in result:
            total["errors"] += 1
            continue
        total["files"] += 1
        for k, v in result.items():
            if k in total:
                total[k] += v

    log.info("=" * 60)
    log.info(f"ИТОГО: {total['files']}/{len(items)} файлов")
    log.info(f"  stocks: {total['stocks']} upserts")
    log.info(f"  ofz:    {total['ofz']} upserts")
    log.info(f"  fx:     {total['fx']} upserts")
    if total["errors"]:
        log.warning(f"  errors: {total['errors']}")


if __name__ == "__main__":
    main()
