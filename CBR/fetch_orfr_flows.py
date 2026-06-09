#!/usr/bin/env python3
"""
Фетчер данных ОРФР Банка России — потоки по категориям участников биржи.

Источник: https://cbr.ru/analytics/finstab/orfr/
Формат: ежемесячные XLSX с фигурами по разделам обзора финансовой стабильности.

Парсит листы «по категориям участников» (вторичные биржевые торги):
  "РА По участникам" — Акции, 7 категорий, кварталы+последние месяцы
  "РО По участникам" — ОФЗ,   7 категорий, кварталы+последние месяцы

⚠️ ЦБ периодически переименовывает листы. До выпуска ORFR_2026-5 (май-2026)
имена были «рис. 32 / рис. 14 / рис. 8»; в мае-2026 ЦБ перешёл на семантические
имена («РА По участникам» = Рынок Акций, «РО По участникам» = Рынок ОФЗ) и
полностью переструктурировал отчёт (43 листа вместо ~10). resolve_sheet_name()
поэтому ищет лист сперва по точному имени, затем по regex заголовка (title_re) —
чтобы следующее переименование не «заморозило» индикатор без алерта.

⚠️ Валюты (старый "рис. 8", 5 категорий по участникам) с мая-2026 БОЛЬШЕ НЕ
публикуются в ОРФР в этом виде — ЦБ заменил единый лист участников на отдельные
одно-рядные серии (РВ Продажи/Покупки нефинансовых, РВ Покупки населением,
РВ продажи 29 экспортеров). instrument_type='fx' заморожен на апреле-2026; новые
данные не загружаются. Если решим возродить fx — remap на новые РВ-листы (другая
методология, потребует правок фронта/методологии).

Каждый новый XLSX содержит весь архив — поэтому ingest стратегия:
взять XLSX, upsert все строки (PK overwrite на повторе, idempotent).

⚠️ РЕЖИМ — РУЧНОЙ. Авто-расписание убрано из main_orchestrator.py:
cbr.ru таймаутит с прод-сервера, а формат отчёта нестабилен. Файл присылается
вручную и грузится через --xlsx. Авто-скачивание (без --xlsx) оставлено как
fallback, но не запускается по расписанию.

Запуск (ручной ингест присланного файла — основной путь):
  python3 -m CBR.fetch_orfr_flows --xlsx /tmp/ORFR_2026-5.xlsx
  python3 -m CBR.fetch_orfr_flows --xlsx FILE --dry-run   # проверить парсинг

CLI flags:
  --xlsx PATH    взять локальный XLSX (основной режим; файл от пользователя)
  --dry-run      не писать в БД, только парсить и вывести summary
  (без --xlsx)   fallback: скачать свежий XLSX с cbr.ru (может таймаутить)
"""

import argparse
import io
import logging
import os
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
import requests
from sqlalchemy import create_engine, text

sys.path.insert(0, str(Path(__file__).parent.parent))
from dotenv import load_dotenv  # noqa: E402

load_dotenv(Path(__file__).parent.parent / ".env")

DB_URL = os.getenv("DB_URL")
LANDING_URL = "https://cbr.ru/analytics/finstab/orfr/"
FILE_URL_TMPL = "https://cbr.ru/Collection/Collection/File/{file_id}/{filename}"

# Сопоставление русских названий месяцев → номер. ЦБ использует Nominative.
MONTH_RU = {
    "Январь": 1, "Февраль": 2, "Март": 3, "Апрель": 4, "Май": 5, "Июнь": 6,
    "Июль": 7, "Август": 8, "Сентябрь": 9, "Октябрь": 10, "Ноябрь": 11, "Декабрь": 12,
}
QUARTER_RU = {
    "I квартал": 1, "II квартал": 2, "III квартал": 3, "IV квартал": 4,
}

# Категории которые ЦБ убрал из методологии — не вставляем в БД.
# КРИТИЧНО: per-instrument! Имена «Нерезиденты», «СЗКО», «Прочие Банки»,
# «Доверительное управление», «Нефинансовые организации» есть И в stocks,
# И в fx. Для fx ЦБ их свернул, для stocks — это основные категории.
# Глобальный set исключал бы их и для акций (баг — концы кварталов теряли
# 5 категорий из 7).
EXCLUDED_BY_TYPE: dict[str, set[str]] = {
    "stocks": {"Дочерние иностранные организации"},  # 2022-2023, свёрнута
    "ofz": {"Дочерние иностранные организации", "Минфин", "Минфин России"},
    "fx": {
        "Банк России (ЦБ РФ)",        # переименовали в «Банк России»
        "Доверительное управление",   # уже не публикуется для валют
        "Нерезиденты",
        "Нефинансовые организации",
        "Прочие Банки",
        "СЗКО",
    },
}

# Конфиг листов: какие парсим и под каким instrument_type сохраняем.
#   sheet     — точное имя листа (актуальное на момент правки)
#   title_re  — regex по заголовку листа (row 0); fallback если ЦБ переименует
#               лист. Заголовок описывает данные и меняется реже имени.
# fx убран: ЦБ прекратил публикацию валют по категориям участников (см. docstring).
SHEETS_CONFIG = [
    {
        "sheet": "РА По участникам",
        "instrument_type": "stocks",
        "label": "Акции",
        "title_re": r"покупки/продажи\s+акций\s+по\s+категориям\s+участник",
    },
    {
        "sheet": "РО По участникам",
        "instrument_type": "ofz",
        "label": "ОФЗ",
        "title_re": r"покупки/продажи\s+ОФЗ\s+по\s+категориям\s+участник",
    },
]

LOG_DIR = Path(__file__).parent / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────────────────
#  1) Поиск свежего XLSX на landing page
# ──────────────────────────────────────────────────────────────────────────────

def find_latest_xlsx_url(timeout: int = 30) -> Optional[tuple[str, str]]:
    """
    Парсит landing page, находит свежий XLSX по (year, month).
    Возвращает (url, basename) или None если ничего не нашлось.

    URL pattern: /Collection/Collection/File/<digits>/ORFR_YYYY-M[_YYYY-M].xlsx
    """
    log.info("Получение списка XLSX с landing page…")
    try:
        resp = requests.get(LANDING_URL, timeout=timeout, headers={
            "User-Agent": "Mozilla/5.0 (compatible; Frame/1.0)"
        })
        resp.raise_for_status()
    except Exception as e:
        log.error(f"Не удалось скачать landing page: {e}")
        return None

    html = resp.text
    # ORFR_YYYY-M.xlsx  или  ORFR_YYYY-M_YYYY-M.xlsx (объединённые выпуски)
    pattern = re.compile(
        r"/Collection/Collection/File/(\d+)/(ORFR_(\d{4})-(\d{1,2})(?:_(\d{4})-(\d{1,2}))?\.xlsx)",
        re.IGNORECASE,
    )
    matches = pattern.findall(html)
    if not matches:
        log.error("Не найдено ни одного XLSX на landing page")
        return None

    # Для объединённых выпусков (ORFR_2025-12_2026-01) берём вторую дату — она новее
    def sort_key(m):
        end_year = int(m[4]) if m[4] else int(m[2])
        end_month = int(m[5]) if m[5] else int(m[3])
        return (end_year, end_month)

    matches.sort(key=sort_key, reverse=True)
    file_id, filename = matches[0][0], matches[0][1]
    url = FILE_URL_TMPL.format(file_id=file_id, filename=filename)
    log.info(f"Свежий XLSX: {filename} (id={file_id})")
    return url, filename


def download_xlsx(url: str, timeout: int = 60) -> bytes:
    log.info(f"Скачивание {url}…")
    resp = requests.get(url, timeout=timeout, headers={
        "User-Agent": "Mozilla/5.0 (compatible; Frame/1.0)"
    })
    resp.raise_for_status()
    log.info(f"Скачано {len(resp.content)/1024:.1f} KB")
    return resp.content


# ──────────────────────────────────────────────────────────────────────────────
#  2) Парсинг XLSX
# ──────────────────────────────────────────────────────────────────────────────

def _quarter_end_date(year: int, q: int) -> date:
    """I → 31.03, II → 30.06, III → 30.09, IV → 31.12."""
    return date(year, q * 3, [31, 30, 30, 31][q - 1])


def _month_end(year: int, month: int) -> date:
    """Последний день месяца."""
    if month == 12:
        return date(year, 12, 31)
    first_next = date(year, month + 1, 1)
    return first_next - timedelta(days=1)


def resolve_sheet_name(wb, cfg: dict) -> Optional[str]:
    """Находит реальное имя листа: сперва точное cfg['sheet'], затем по regex
    заголовка (cfg['title_re']) в первых 3 строках.

    ЦБ периодически переименовывает листы (рис.N → семантические имена в
    мае-2026). Точное имя — быстрый путь; title_re — устойчивость к будущим
    переименованиям, чтобы фетчер не «замораживал» индикатор молча.
    """
    if cfg["sheet"] in wb.sheetnames:
        return cfg["sheet"]
    title_re = cfg.get("title_re")
    if not title_re:
        return None
    pat = re.compile(title_re, re.IGNORECASE)
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for cell in row:
                if isinstance(cell, str) and pat.search(cell):
                    log.warning(
                        f"Лист {cfg['sheet']!r} не найден — matched по заголовку: {name!r}"
                    )
                    return name
    return None


def parse_sheet(wb, sheet_name: str, source_file: str, instrument_type: str) -> Iterable[dict]:
    """
    Yields {period_year, period_label, period_kind, period_end_date, category, value}.

    Структура листа (одинаковая у всех 3):
      Row 1: title (skip)
      Row 2-3: meta/source (skip)
      Row 4: empty
      Row 5: header — "Годы" | "Месяцы/кварталы" | <cat1> | <cat2> | ... | <catN>
      Row 6+: data — год (sometimes empty) | период | val1 | val2 | ...

    Год повторяется только в первой row блока для группы → нужно forward-fill.
    """
    if sheet_name not in wb.sheetnames:
        log.warning(f"Лист {sheet_name!r} не найден — пропуск")
        return

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Найти header (row с "Годы"/"Год" в первой колонке)
    header_idx = None
    for idx, r in enumerate(rows):
        if r and r[0] and isinstance(r[0], str) and r[0].strip() in ("Годы", "Год"):
            header_idx = idx
            break
    if header_idx is None:
        log.error(f"{sheet_name}: header (строка с 'Годы') не найден")
        return

    header = rows[header_idx]
    # Категории — колонки начиная с index 2 (0=Год, 1=Период, 2+=категории)
    categories = []
    for i, cell in enumerate(header[2:], start=2):
        if cell and isinstance(cell, str) and cell.strip():
            categories.append((i, cell.strip()))

    log.info(f"{sheet_name}: {len(categories)} категорий: {[c for _, c in categories]}")

    current_year: Optional[int] = None
    parsed = 0

    for r in rows[header_idx + 1:]:
        if not r or all(c is None or c == "" for c in r[:2]):
            continue
        year_cell, period_cell = r[0], r[1]
        # Forward-fill year
        if year_cell is not None and year_cell != "":
            try:
                current_year = int(year_cell)
            except (ValueError, TypeError):
                continue
        if current_year is None or period_cell is None or not isinstance(period_cell, str):
            continue

        plabel = period_cell.strip()
        # Определяем kind и end_date.
        # КВАРТАЛЫ ПРОПУСКАЕМ: индикатор работает только с месячными данными.
        # Квартальная строка имеет тот же period_end_date что последний месяц
        # квартала (31.03/30.06/30.09/31.12) → через ON CONFLICT затирала бы
        # корректную месячную запись. Парсим только month.
        if plabel in QUARTER_RU:
            continue
        elif plabel in MONTH_RU:
            kind = "month"
            end_dt = _month_end(current_year, MONTH_RU[plabel])
        else:
            log.debug(f"{sheet_name}: неизвестный период {plabel!r} в году {current_year} — пропуск")
            continue

        excluded = EXCLUDED_BY_TYPE.get(instrument_type, set())
        for col_idx, cat in categories:
            if cat in excluded:
                continue
            val = r[col_idx] if col_idx < len(r) else None
            if val is None:
                continue
            try:
                fval = float(val)
            except (ValueError, TypeError):
                continue
            yield {
                "period_year": current_year,
                "period_label": plabel,
                "period_kind": kind,
                "period_end_date": end_dt,
                "category": cat,
                "value": fval,
                "source_file": source_file,
            }
            parsed += 1

    log.info(f"{sheet_name}: распарсено {parsed} (период × категория) точек")


# ──────────────────────────────────────────────────────────────────────────────
#  3) Upsert в БД
# ──────────────────────────────────────────────────────────────────────────────

UPSERT_SQL = text("""
    INSERT INTO cbr_flows (
        instrument_type, period_year, period_label, period_kind,
        period_end_date, category, value, source_file, updated_at
    ) VALUES (
        :instrument_type, :period_year, :period_label, :period_kind,
        :period_end_date, :category, :value, :source_file, NOW()
    )
    ON CONFLICT (instrument_type, period_end_date, category)
    DO UPDATE SET
        value = EXCLUDED.value,
        period_label = EXCLUDED.period_label,
        period_kind = EXCLUDED.period_kind,
        period_year = EXCLUDED.period_year,
        source_file = EXCLUDED.source_file,
        updated_at = NOW()
""")


def upsert_rows(engine, instrument_type: str, rows: list[dict]) -> int:
    if not rows:
        return 0
    # FX в XLSX ЦБ использует ОБРАТНУЮ конвенцию знаков относительно акций/ОФЗ:
    # лист «рис. 8» имеет заголовок «Нетто-покупка(-) / нетто-продажа (+)
    # валюты за рубли». То есть минус = покупка, плюс = продажа.
    # Stocks/OFZ (рис. 32/14) используют стандартно: + = покупка, − = продажа.
    # Чтобы унифицировать БД (везде «+ = покупка»), инвертируем fx values.
    # После этого вся pipeline (API, frontend, tooltip) trapping одинаково.
    if instrument_type == "fx":
        rows = [{**r, "value": -r["value"]} for r in rows]
    payload = [{**r, "instrument_type": instrument_type} for r in rows]
    with engine.begin() as conn:
        conn.execute(UPSERT_SQL, payload)
    return len(payload)


# ──────────────────────────────────────────────────────────────────────────────
#  4) Главный pipeline
# ──────────────────────────────────────────────────────────────────────────────

def run(xlsx_path: Optional[str] = None, dry_run: bool = False) -> int:
    """Возвращает количество upsert'нутых rows (0 если dry-run)."""

    # 1) Получаем XLSX bytes
    if xlsx_path:
        log.info(f"Использую локальный XLSX: {xlsx_path}")
        with open(xlsx_path, "rb") as f:
            xlsx_bytes = f.read()
        source_file = Path(xlsx_path).stem
    else:
        url_info = find_latest_xlsx_url()
        if not url_info:
            log.critical("Не удалось найти XLSX на cbr.ru")
            return 0
        url, filename = url_info
        xlsx_bytes = download_xlsx(url)
        source_file = filename.rsplit(".", 1)[0]

    # 2) Загружаем workbook
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=False)

    # 3) Парсим листы
    all_rows: dict[str, list[dict]] = {}
    for cfg in SHEETS_CONFIG:
        sheet_name = resolve_sheet_name(wb, cfg)
        if not sheet_name:
            log.warning(
                f"  → {cfg['label']:8s} ({cfg['instrument_type']}): лист не найден "
                f"(ни по имени {cfg['sheet']!r}, ни по заголовку) — пропуск"
            )
            all_rows[cfg["instrument_type"]] = []
            continue
        rows = list(parse_sheet(wb, sheet_name, source_file, cfg["instrument_type"]))
        all_rows[cfg["instrument_type"]] = rows
        log.info(f"  → {cfg['label']:8s} ({cfg['instrument_type']}): {len(rows)} rows")

    if dry_run:
        log.info("DRY-RUN — БД не трогаем")
        for itype, rows in all_rows.items():
            if rows:
                latest = max(rows, key=lambda r: r["period_end_date"])
                log.info(f"  {itype}: latest={latest['period_end_date']} {latest['period_label']}")
        return 0

    # 4) Upsert
    if not DB_URL:
        log.critical("DB_URL не задан в .env!")
        return 0

    engine = create_engine(DB_URL, connect_args={"ssl_context": False})
    total = 0
    for itype, rows in all_rows.items():
        if not rows:
            continue
        n = upsert_rows(engine, itype, rows)
        log.info(f"  ✓ {itype}: upsert {n} rows")
        total += n
    log.info(f"Готово: всего upsert'нуто {total} rows")
    return total


def main():
    p = argparse.ArgumentParser(description="Фетчер ОРФР ЦБ — потоки участников биржи")
    p.add_argument("--xlsx", help="Локальный путь к XLSX (вместо скачивания)")
    p.add_argument("--dry-run", action="store_true", help="Не писать в БД")
    p.add_argument("--once", action="store_true", help="Single-shot (alias, и так по умолчанию)")
    args = p.parse_args()

    try:
        n = run(xlsx_path=args.xlsx, dry_run=args.dry_run)
        sys.exit(0 if n > 0 or args.dry_run else 1)
    except Exception as e:
        log.exception(f"Fatal error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
