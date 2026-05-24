"""
CSV export helper — общий компонент для всех индикатор-export-endpoint'ов.

Архитектура:
  - StreamingResponse — не держим весь dataset в памяти (некоторые export'ы
    могут быть до 100k строк).
  - UTF-8 BOM (﻿) перед заголовком — Excel правильно opens кириллицу
    без манипуляций с кодировкой.
  - Content-Disposition: attachment — браузер сразу триггерит download.

При выборе нескольких слоёв (`?layers=A,B`) — endpoint делает ZIP с
отдельными CSV-файлами per layer.
"""
import csv
import io
import zipfile
from typing import Iterable, Iterator

from fastapi.responses import StreamingResponse, Response


def _bom() -> str:
    """UTF-8 BOM — Excel-friendly кириллица."""
    return "﻿"


def csv_streaming_response(
    rows: Iterable[dict],
    fieldnames: list[str],
    filename: str,
) -> StreamingResponse:
    """
    Стримит CSV-файл клиенту.

    Args:
        rows: iterable of dicts. Каждый dict должен содержать ключи из fieldnames.
              Отсутствующие — пустая ячейка.
        fieldnames: порядок колонок и список заголовков.
        filename: имя файла для browser download (без пути).

    Returns:
        FastAPI StreamingResponse с правильными headers.
    """
    def generator() -> Iterator[str]:
        # Header line + UTF-8 BOM (один раз, в начале).
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        yield _bom() + buffer.getvalue()
        buffer.seek(0)
        buffer.truncate(0)

        # Data rows — каждый flushим отдельно.
        for row in rows:
            writer.writerow(row)
            yield buffer.getvalue()
            buffer.seek(0)
            buffer.truncate(0)

    return StreamingResponse(
        generator(),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            # Не кэшируем — данные актуальные на момент запроса.
            "Cache-Control": "no-store",
        },
    )


def _csv_blob(rows: Iterable[dict], fieldnames: list[str]) -> str:
    """Собирает CSV целиком в str (для ZIP packaging — нужен seek)."""
    buffer = io.StringIO()
    buffer.write(_bom())
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue()


def xlsx_response(
    sheets: dict[str, tuple[Iterable[dict], list[str]]],
    filename: str,
) -> Response:
    """
    Упаковывает несколько rowsets в XLSX (один файл, multiple sheets).

    Преимущество vs ZIP: один файл, Excel-нативная навигация по табам
    (sheet'ам). Размер сравним с ZIP-of-CSVs.

    Args:
        sheets: {sheet_name: (rows, fieldnames)}. Sheet name max 31 char
                (Excel limit), запрещены чары \\ / ? * [ ].
        filename: имя .xlsx файла для download.
    """
    from openpyxl import Workbook

    wb = Workbook()
    # Удаляем default sheet который Workbook() создаёт автоматически.
    wb.remove(wb.active)

    for raw_name, (rows, fieldnames) in sheets.items():
        # Sanitize sheet name под Excel-ограничения.
        safe_name = _sanitize_sheet_name(raw_name)
        ws = wb.create_sheet(title=safe_name)
        # Header row.
        ws.append(fieldnames)
        for row in rows:
            ws.append([row.get(f) for f in fieldnames])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )


def _sanitize_sheet_name(name: str) -> str:
    """Excel запрещает / \\ ? * [ ] в sheet name + max 31 char."""
    sanitized = name
    for ch in r"/\?*[]:":
        sanitized = sanitized.replace(ch, "_")
    # Strip .csv extension если осталось из ZIP-naming convention.
    if sanitized.endswith(".csv"):
        sanitized = sanitized[:-4]
    return sanitized[:31]


def zip_response(
    files: dict[str, tuple[Iterable[dict], list[str]]],
    filename: str,
) -> Response:
    """
    Упаковывает несколько CSV-блобов в один ZIP.

    Args:
        files: {filename.csv: (rows, fieldnames)}.
        filename: имя ZIP файла для browser download.

    Streaming не использован — CSV-блоб должен быть весь в памяти для
    ZIP-archive (zipfile требует seek). Для наших размеров (<10MB total)
    приемлемо.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for name, (rows, fieldnames) in files.items():
            zf.writestr(name, _csv_blob(rows, fieldnames))
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )
