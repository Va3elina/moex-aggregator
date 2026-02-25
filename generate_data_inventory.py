#!/usr/bin/env python3
"""
Генерация Excel-инвентаризации данных БД moex_db для договора с МосБиржей.
"""
import sys
import io
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import os
from datetime import datetime
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()
DB_URL = os.getenv("DB_URL")
engine = create_engine(DB_URL)

wb = Workbook()

# === Стили ===
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
subheader_font = Font(bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center')


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def style_row(ws, row, cols, fill=None):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if fill:
            cell.fill = fill
            cell.font = subheader_font


def auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                max_line = max(len(l) for l in lines)
                max_len = max(max_len, max_line)
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


# =========================================================================
# ЛИСТ 1: ФЬЮЧЕРСЫ — сводная таблица
# =========================================================================
ws1 = wb.active
ws1.title = "Фьючерсы"
ws1.sheet_properties.tabColor = "2F5496"

headers1 = [
    "Тикер (sectype)", "Название", "Группа",
    "Свечи 5м\n(строк)", "Свечи 5м\nс", "Свечи 5м\nпо",
    "Свечи 60м\n(строк)", "Свечи 60м\nс", "Свечи 60м\nпо",
    "Свечи 24ч\n(строк)", "Свечи 24ч\nс", "Свечи 24ч\nпо",
    "ОИ 5м\n(строк)", "ОИ 5м\nс", "ОИ 5м\nпо",
    "ОИ 60м\n(строк)", "ОИ 60м\nс", "ОИ 60м\nпо",
    "ОИ 24ч\n(строк)", "ОИ 24ч\nс", "ОИ 24ч\nпо",
    "Источник свечей", "Источник ОИ"
]

for i, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, 1, len(headers1))
ws1.row_dimensions[1].height = 40

with engine.connect() as conn:
    # Instruments futures (unique by sectype)
    instruments = conn.execute(text("""
        SELECT DISTINCT sectype, name, "group"
        FROM instruments
        WHERE type = 'futures'
        ORDER BY sectype
    """)).fetchall()

    # Candles by sec_id (sec_id maps to instrument prefix for futures)
    candles_raw = conn.execute(text("""
        SELECT sec_id, interval, COUNT(*) as rows,
               MIN(begin_time)::date::text, MAX(begin_time)::date::text
        FROM candles WHERE type = 'futures' AND sec_id IS NOT NULL
        GROUP BY sec_id, interval
    """)).fetchall()

    # OI by sectype
    oi_raw = conn.execute(text("""
        SELECT sectype, interval, COUNT(*) as rows,
               MIN(tradedate)::text, MAX(tradedate)::text
        FROM open_interest
        GROUP BY sectype, interval
    """)).fetchall()

# Build candles lookup: {sec_id: {interval: (rows, min, max)}}
candles = {}
for r in candles_raw:
    sec_id, interval, rows, min_dt, max_dt = r
    candles.setdefault(sec_id, {})[interval] = (rows, min_dt, max_dt)

# Build OI lookup: {sectype: {interval: (rows, min, max)}}
oi = {}
for r in oi_raw:
    sectype, interval, rows, min_dt, max_dt = r
    oi.setdefault(sectype, {})[interval] = (rows, min_dt, max_dt)

# Collect unique sectypes (some instruments share sectype)
seen_sectypes = set()
row_num = 2

for instr in instruments:
    sectype, name, group = instr
    if sectype in seen_sectypes:
        continue
    seen_sectypes.add(sectype)

    # Find matching sec_ids for this sectype in candles
    # sec_id for futures is like "SiH", "SiM" etc., sectype is "Si"
    # We aggregate across all sec_ids that start with sectype
    candle_data = {}
    for sec_id, intervals in candles.items():
        # Match: sec_id starts with sectype and the rest is 0-1 letter (month code)
        if sec_id == sectype or (sec_id.startswith(sectype) and len(sec_id) <= len(sectype) + 1):
            for interval, (rows, min_dt, max_dt) in intervals.items():
                if interval not in candle_data:
                    candle_data[interval] = [0, '9999-99-99', '0000-00-00']
                candle_data[interval][0] += rows
                candle_data[interval][1] = min(candle_data[interval][1], min_dt)
                candle_data[interval][2] = max(candle_data[interval][2], max_dt)

    oi_data = oi.get(sectype, {})

    def get_candle(interval):
        d = candle_data.get(interval)
        if d and d[0] > 0:
            return d[0], d[1], d[2]
        return '', '', ''

    def get_oi(interval):
        d = oi_data.get(interval)
        if d:
            return d
        return '', '', ''

    c5 = get_candle(5)
    c60 = get_candle(60)
    c24 = get_candle(24)
    o5 = get_oi(5)
    o60 = get_oi(60)
    o24 = get_oi(24)

    # Determine sources
    candle_src = "MOEX Algopack API" if any(v for v in [c5[0], c60[0], c24[0]] if v) else ""
    oi_sources = []
    if o5[0]:
        oi_sources.append("Algopack API (5м)")
    if o60[0]:
        oi_sources.append("Агрегация (60м)")
    if o24[0]:
        if sectype in [s for s, _ , _ in instruments]:
            oi_sources.append("ISS MOEX (24ч)")
    oi_src = "; ".join(oi_sources)

    ws1.cell(row=row_num, column=1, value=sectype)
    ws1.cell(row=row_num, column=2, value=name)
    ws1.cell(row=row_num, column=3, value=group or '')

    ws1.cell(row=row_num, column=4, value=c5[0] if c5[0] else '')
    ws1.cell(row=row_num, column=5, value=c5[1] if c5[1] else '')
    ws1.cell(row=row_num, column=6, value=c5[2] if c5[2] else '')
    ws1.cell(row=row_num, column=7, value=c60[0] if c60[0] else '')
    ws1.cell(row=row_num, column=8, value=c60[1] if c60[1] else '')
    ws1.cell(row=row_num, column=9, value=c60[2] if c60[2] else '')
    ws1.cell(row=row_num, column=10, value=c24[0] if c24[0] else '')
    ws1.cell(row=row_num, column=11, value=c24[1] if c24[1] else '')
    ws1.cell(row=row_num, column=12, value=c24[2] if c24[2] else '')

    ws1.cell(row=row_num, column=13, value=o5[0] if o5[0] else '')
    ws1.cell(row=row_num, column=14, value=o5[1] if o5[1] else '')
    ws1.cell(row=row_num, column=15, value=o5[2] if o5[2] else '')
    ws1.cell(row=row_num, column=16, value=o60[0] if o60[0] else '')
    ws1.cell(row=row_num, column=17, value=o60[1] if o60[1] else '')
    ws1.cell(row=row_num, column=18, value=o60[2] if o60[2] else '')
    ws1.cell(row=row_num, column=19, value=o24[0] if o24[0] else '')
    ws1.cell(row=row_num, column=20, value=o24[1] if o24[1] else '')
    ws1.cell(row=row_num, column=21, value=o24[2] if o24[2] else '')

    ws1.cell(row=row_num, column=22, value=candle_src)
    ws1.cell(row=row_num, column=23, value=oi_src)

    style_row(ws1, row_num, len(headers1))
    row_num += 1

ws1.freeze_panes = 'A2'
auto_width(ws1, min_width=8, max_width=22)

# =========================================================================
# ЛИСТ 2: АКЦИИ
# =========================================================================
ws2 = wb.create_sheet("Акции")
ws2.sheet_properties.tabColor = "548235"

headers2 = [
    "Тикер", "Название",
    "Свечи 5м\n(строк)", "Свечи 5м\nс", "Свечи 5м\nпо",
    "Свечи 60м\n(строк)", "Свечи 60м\nс", "Свечи 60м\nпо",
    "Свечи 24ч\n(строк)", "Свечи 24ч\nс", "Свечи 24ч\nпо",
    "Источник"
]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, len(headers2))
ws2.row_dimensions[1].height = 40

with engine.connect() as conn:
    stocks = conn.execute(text("""
        SELECT sec_id, name FROM instruments WHERE type = 'stock' ORDER BY sec_id
    """)).fetchall()

    stock_candles = conn.execute(text("""
        SELECT sec_id, interval, COUNT(*),
               MIN(begin_time)::date::text, MAX(begin_time)::date::text
        FROM candles WHERE type = 'stock' AND sec_id IS NOT NULL
        GROUP BY sec_id, interval
    """)).fetchall()

sc = {}
for r in stock_candles:
    sc.setdefault(r[0], {})[r[1]] = (r[2], r[3], r[4])

row_num = 2
for sec_id, name in stocks:
    data = sc.get(sec_id, {})
    c5 = data.get(5, ('', '', ''))
    c60 = data.get(60, ('', '', ''))
    c24 = data.get(24, ('', '', ''))

    ws2.cell(row=row_num, column=1, value=sec_id)
    ws2.cell(row=row_num, column=2, value=name)
    ws2.cell(row=row_num, column=3, value=c5[0] if c5[0] else '')
    ws2.cell(row=row_num, column=4, value=c5[1] if c5[1] else '')
    ws2.cell(row=row_num, column=5, value=c5[2] if c5[2] else '')
    ws2.cell(row=row_num, column=6, value=c60[0] if c60[0] else '')
    ws2.cell(row=row_num, column=7, value=c60[1] if c60[1] else '')
    ws2.cell(row=row_num, column=8, value=c60[2] if c60[2] else '')
    ws2.cell(row=row_num, column=9, value=c24[0] if c24[0] else '')
    ws2.cell(row=row_num, column=10, value=c24[1] if c24[1] else '')
    ws2.cell(row=row_num, column=11, value=c24[2] if c24[2] else '')
    ws2.cell(row=row_num, column=12, value="MOEX Algopack API")

    style_row(ws2, row_num, len(headers2))
    row_num += 1

ws2.freeze_panes = 'A2'
auto_width(ws2, min_width=8, max_width=22)

# =========================================================================
# ЛИСТ 3: Индексы
# =========================================================================
ws3 = wb.create_sheet("Индексы")
ws3.sheet_properties.tabColor = "BF8F00"

headers3 = [
    "Индекс (secid)", "Название",
    "Строк", "Период с", "Период по",
    "Данные", "Источник"
]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, len(headers3))

with engine.connect() as conn:
    idx_data = conn.execute(text("""
        SELECT i.secid, i.name,
               COUNT(d.id), MIN(d.trade_date)::text, MAX(d.trade_date)::text
        FROM indices i
        LEFT JOIN index_data d ON i.secid = d.secid
        GROUP BY i.secid, i.name
        ORDER BY i.secid
    """)).fetchall()

row_num = 2
for r in idx_data:
    ws3.cell(row=row_num, column=1, value=r[0])
    ws3.cell(row=row_num, column=2, value=r[1])
    ws3.cell(row=row_num, column=3, value=r[2])
    ws3.cell(row=row_num, column=4, value=r[3] or '')
    ws3.cell(row=row_num, column=5, value=r[4] or '')
    ws3.cell(row=row_num, column=6, value="OHLC, Value, Capitalization (дневные)")
    ws3.cell(row=row_num, column=7, value="ISS MOEX History API")
    style_row(ws3, row_num, len(headers3))
    row_num += 1

ws3.freeze_panes = 'A2'
auto_width(ws3)

# =========================================================================
# ЛИСТ 4: БПИФ (Фонды)
# =========================================================================
ws4 = wb.create_sheet("БПИФ")
ws4.sheet_properties.tabColor = "538135"

headers4 = [
    "Тикер", "Название", "УК", "Категория",
    "Строк", "Период с", "Период по",
    "Данные", "Источник"
]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, len(headers4))

with engine.connect() as conn:
    fund_rows = conn.execute(text("""
        SELECT f.ticker, f.name, f.uk, f.category,
               COUNT(d.id), MIN(d.trade_date)::text, MAX(d.trade_date)::text
        FROM funds f
        LEFT JOIN fund_data d ON f.fund_id = d.fund_id
        GROUP BY f.ticker, f.name, f.uk, f.category
        ORDER BY f.category, f.ticker
    """)).fetchall()

row_num = 2
for r in fund_rows:
    cat_map = {
        'money_market': 'Денежный рынок',
        'stocks': 'Акции',
        'bonds': 'Облигации',
        'gold': 'Золото'
    }
    ws4.cell(row=row_num, column=1, value=r[0])
    ws4.cell(row=row_num, column=2, value=r[1])
    ws4.cell(row=row_num, column=3, value=r[2])
    ws4.cell(row=row_num, column=4, value=cat_map.get(r[3], r[3]))
    ws4.cell(row=row_num, column=5, value=r[4])
    ws4.cell(row=row_num, column=6, value=r[5] or '')
    ws4.cell(row=row_num, column=7, value=r[6] or '')
    ws4.cell(row=row_num, column=8, value="Цена пая (PAY), СЧА (NAV) (дневные)")
    ws4.cell(row=row_num, column=9, value="ISS MOEX (shares TQTF + index iNAV)")
    style_row(ws4, row_num, len(headers4))
    row_num += 1

ws4.freeze_panes = 'A2'
auto_width(ws4)

# =========================================================================
# ЛИСТ 5: Макроданные
# =========================================================================
ws5 = wb.create_sheet("Макроданные")
ws5.sheet_properties.tabColor = "C00000"

headers5 = [
    "Индикатор", "Название", "Частота",
    "Строк", "Период с", "Период по",
    "Источник данных", "Данные с MOEX?"
]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, len(headers5))

macro_items = [
    ("M2_MONTHLY", "Денежная масса M2 (млрд руб)", "Ежемесячно", "CBR_XLSX", "ЦБ РФ (cbr.ru)", "НЕТ"),
    ("GDP_QUARTERLY", "ВВП России (млрд руб)", "Ежеквартально", "ROSSTAT_XLSX", "Росстат (ручной импорт)", "НЕТ"),
    ("MARKET_CAP_TOTAL", "Капитализация фонд. рынка РФ (млрд руб)", "Ежедневно", "SMARTLAB + GURUFOCUS", "SmartLab + GuruFocus/WebArchive", "НЕТ"),
]

with engine.connect() as conn:
    macro_stats = conn.execute(text("""
        SELECT indicator, COUNT(*), MIN(period_date)::text, MAX(period_date)::text
        FROM macro_data GROUP BY indicator ORDER BY indicator
    """)).fetchall()

ms = {r[0]: (r[1], r[2], r[3]) for r in macro_stats}

row_num = 2
for ind, name, freq, src_key, src_desc, is_moex in macro_items:
    stats = ms.get(ind, (0, '', ''))
    ws5.cell(row=row_num, column=1, value=ind)
    ws5.cell(row=row_num, column=2, value=name)
    ws5.cell(row=row_num, column=3, value=freq)
    ws5.cell(row=row_num, column=4, value=stats[0])
    ws5.cell(row=row_num, column=5, value=stats[1] or '')
    ws5.cell(row=row_num, column=6, value=stats[2] or '')
    ws5.cell(row=row_num, column=7, value=src_desc)
    ws5.cell(row=row_num, column=8, value=is_moex)
    style_row(ws5, row_num, len(headers5))
    row_num += 1

ws5.freeze_panes = 'A2'
auto_width(ws5)

# =========================================================================
# ЛИСТ 6: Сводка (общая статистика)
# =========================================================================
ws6 = wb.create_sheet("Сводка")
ws6.sheet_properties.tabColor = "7030A0"
# Move to first position
wb.move_sheet(ws6, offset=-5)

headers6 = ["Таблица / Набор данных", "Описание", "Строк", "Размер", "Период с", "Период по", "Источник данных", "API endpoint"]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=i, value=h)
style_header(ws6, 1, len(headers6))

summary_rows = [
    ("candles (фьючерсы, 5м)", "OHLCV свечи фьючерсов, 5 мин (агрегация из 1м)", "10,493,859", "",
     "2021-11-08", "2026-01-30", "MOEX Algopack API", "apim.moex.com/iss/.../candles.json"),
    ("candles (фьючерсы, 60м)", "OHLCV свечи фьючерсов, часовые", "2,881,857", "",
     "2011-12-19", "2026-01-30", "MOEX Algopack API", "apim.moex.com/iss/.../candles.json"),
    ("candles (фьючерсы, 24ч)", "OHLCV свечи фьючерсов, дневные", "362,035", "",
     "2002-03-25", "2026-01-30", "MOEX Algopack API", "apim.moex.com/iss/.../candles.json"),
    ("candles (акции, 5м)", "OHLCV свечи акций, 5 мин", "2,027,080", "",
     "2021-11-09", "2026-01-30", "MOEX Algopack API", "apim.moex.com/iss/.../candles.json"),
    ("candles (акции, 60м)", "OHLCV свечи акций, часовые", "1,465,210", "",
     "2011-11-17", "2026-01-30", "MOEX Algopack API", "apim.moex.com/iss/.../candles.json"),
    ("candles (акции, 24ч)", "OHLCV свечи акций, дневные", "154,186", "",
     "2007-01-09", "2026-01-30", "MOEX Algopack API", "apim.moex.com/iss/.../candles.json"),
    ("open_interest (5м)", "ОИ 5-мин, ФИЗ+ЮР, 65 тикеров", "21,631,522", "",
     "2020-01-03", "2026-02-07", "MOEX Algopack API", "apim.moex.com/iss/analyticalproducts/futoi/..."),
    ("open_interest (60м)", "ОИ часовой (агрегация из 5м)", "1,821,394", "",
     "2020-01-03", "2026-02-07", "Внутренняя агрегация", "—"),
    ("open_interest (24ч)", "ОИ дневной, ФИЗ+ЮР, 133 инстр.", "353,720", "",
     "2012-10-01", "2026-01-30", "ISS MOEX (бесплатный)", "iss.moex.com/iss/statistics/.../openpositions/..."),
    ("index_data", "Дневные данные индексов (5 шт)", "13,414", "4.2 MB",
     "2005-01-11", "2026-02-10", "ISS MOEX History API", "iss.moex.com/iss/history/..."),
    ("fund_data", "Дневные данные БПИФ (15 фондов)", "14,685", "2.5 MB",
     "2020-01-03", "2026-02-09", "ISS MOEX History API", "iss.moex.com/iss/history/..."),
    ("macro_data (M2)", "Денежная масса M2, ежемесячно", "397", "",
     "1993-01-31", "2026-01-31", "ЦБ РФ (НЕ MOEX)", "cbr.ru XLSX"),
    ("macro_data (GDP)", "ВВП, ежеквартально", "123", "",
     "1995-03-31", "2025-09-30", "Росстат (НЕ MOEX)", "ручной импорт XLSX"),
    ("macro_data (Market Cap)", "Капитализация рынка РФ", "322", "",
     "1997-09-20", "2026-02-17", "SmartLab + GuruFocus (НЕ MOEX)", "smart-lab.ru HTML парсинг"),
    ("instruments", "Справочник инструментов", "391", "200 KB",
     "—", "—", "Ручное заполнение", "—"),
]

row_num = 2
for r in summary_rows:
    for i, val in enumerate(r, 1):
        ws6.cell(row=row_num, column=i, value=val)
    style_row(ws6, row_num, len(headers6))
    row_num += 1

# Total row
row_num += 1
ws6.cell(row=row_num, column=1, value="ИТОГО")
ws6.cell(row=row_num, column=2, value="Вся база данных")
ws6.cell(row=row_num, column=3, value="~41,200,000")
ws6.cell(row=row_num, column=4, value="12 GB")
for col in range(1, len(headers6) + 1):
    cell = ws6.cell(row=row_num, column=col)
    cell.font = Font(bold=True, size=11)
    cell.fill = subheader_fill
    cell.border = thin_border

ws6.freeze_panes = 'A2'
auto_width(ws6, min_width=12, max_width=55)

# =========================================================================
# СОХРАНЕНИЕ
# =========================================================================
output_path = os.path.join(os.path.dirname(__file__), "MOEX_Data_Inventory.xlsx")
wb.save(output_path)
print(f"Excel saved: {output_path}")
print("Done!")
