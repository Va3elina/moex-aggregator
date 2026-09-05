#!/usr/bin/env python3
"""
Автоматическое обновление макроэкономических данных (M2, GDP).

Источники:
- M2: ЦБ РФ XLSX
  https://www.cbr.ru/vfs/statistics/credit_statistics/monetary_agg.xlsx
- GDP: Росстат, авто-discover ссылки на VVP_kvartal_s_*.xlsx со страницы
  https://rosstat.gov.ru/statistics/accounts (имя файла меняется ежегодно)

Расписание: ежедневно (M2 ежемесячно, GDP — раз в квартал, но мы каждый день
проверяем не появилось ли свежее).

Использование:
  python Macro/fetch_macro_realtime.py --once             # Обновить M2 + GDP
  python Macro/fetch_macro_realtime.py --once --force     # Полная перезагрузка
  python Macro/fetch_macro_realtime.py --import-gdp FILE  # Ручной импорт GDP (fallback)
  python Macro/fetch_macro_realtime.py --check            # Проверка целостности
"""

import asyncio
import json
import argparse
import logging
import subprocess
import sys
import os
import tempfile
import urllib.request
from pathlib import Path
from datetime import datetime, date, timedelta
from calendar import monthrange
from typing import List, Tuple, Optional, Dict

from dotenv import load_dotenv
from sqlalchemy import create_engine, text

# === Пути ===
BASE_DIR = Path(__file__).parent
PROJECT_DIR = BASE_DIR.parent

# Оркестратор запускает скрипт с cwd=Macro/, поэтому корень репо в sys.path
# сам не попадает. Нужен и для api.ru_tls (TLS-контекст Росстата ниже), и для
# moex_calendar.
sys.path.insert(0, str(PROJECT_DIR))

load_dotenv(PROJECT_DIR / ".env")
DB_URL = os.getenv("DB_URL")

LOG_DIR = BASE_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

# === URL источников ===
CBR_M2_URL = "https://www.cbr.ru/vfs/statistics/credit_statistics/monetary_agg.xlsx"

# Росстат: индексная страница «Национальные счета» (стабильна, имя файла меняется
# ежегодно: VVP_kvartal_s_1995-2025.xlsx → VVP_kvartal_s_1995-2026.xlsx и т.д.).
# Поэтому URL файла извлекается из HTML страницы регуляркой.
ROSSTAT_GDP_PAGE_URL = "https://rosstat.gov.ru/statistics/accounts"
ROSSTAT_BASE_URL = "https://rosstat.gov.ru"
# Сертификат rosstat.gov.ru подписан Russian Trusted Root CA (Минцифры),
# которого нет в certifi, и сам Росстат не присылает промежуточный сертификат.
# Цепочку достраиваем своим бандлом — см. api/ru_tls.py::rosstat_ssl_context.
# До 30.07.2026 здесь стоял CERT_NONE, то есть данные ВВП принимались вообще
# без проверки сертификата.
import re as _re
import ssl as _ssl
from api.ru_tls import rosstat_ssl_context as _ru_tls_rosstat_context


def _rosstat_ssl_context() -> _ssl.SSLContext:
    """Контекст с полной проверкой цепочки Минцифры.

    Тонкая обёртка: имя приватное, но на него ссылаются рунбуки ручного
    импорта ВВП (скилл moex-macro-refresh), поэтому сохраняем как есть.
    """
    return _ru_tls_rosstat_context()


def _tls_hint(exc: Exception) -> str:
    """Подсказка, если упала именно проверка сертификата Росстата.

    Сбой ВВП не красит пайплайн (GDP-фетч некритичен), так что единственный
    след — эта строка в логе. Пусть она сразу говорит, что чинить.
    """
    if "CERTIFICATE_VERIFY_FAILED" not in str(exc):
        return ""
    return (
        " — не сошлась цепочка Минцифры. Вероятнее всего Росстату выписали "
        "лист под новым Sub CA: обновить certs/russian_trusted_sub_ca.pem "
        "(порядок действий — в шапке этого файла), проверка — "
        "python scripts/check_ru_tls.py"
    )

# === Лимиты валидации ===
M2_MIN_VALUE = 1            # M2 min (млрд руб) — в 1993 M2 было ~6.5 млрд руб
M2_MAX_VALUE = 200_000_000  # M2 max (200 трлн руб) — запас на будущее
M2_MAX_JUMP_PCT = 30        # Максимальный скачок между соседними месяцами (%)
M2_MIN_DATE = date(1993, 1, 1)   # M2 доступна с 1993

GDP_MIN_VALUE = 100         # GDP min (млрд руб)
GDP_MAX_VALUE = 100_000_000 # GDP max (100 трлн руб за квартал)
GDP_MAX_JUMP_PCT = 50       # Максимальный скачок между кварталами (%)
GDP_MIN_DATE = date(1995, 1, 1)  # GDP доступна с 1995

# === Календарь MOEX ===  (PROJECT_DIR уже в sys.path, см. блок «Пути»)
try:
    from moex_calendar import get_moscow_time, is_trading_day
except ImportError:
    def get_moscow_time():
        return datetime.utcnow() + timedelta(hours=3)
    def is_trading_day(check_date=None):
        check = check_date or get_moscow_time().date()
        if check.weekday() in [5, 6]:
            return False, "Выходной"
        return True, "Торговый день"


# ======================================================================
#                         ЛОГИРОВАНИЕ
# ======================================================================

def setup_logging():
    """Настройка логирования: консоль + файл всех логов + файл ошибок."""
    detailed_fmt = logging.Formatter(
        '%(asctime)s | %(levelname)-8s | %(funcName)s:%(lineno)d | %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    console_fmt = logging.Formatter(
        '%(asctime)s | %(levelname)-8s | %(message)s',
        datefmt='%H:%M:%S'
    )

    root = logging.getLogger()
    root.setLevel(logging.DEBUG)
    root.handlers.clear()

    # Консоль (INFO+)
    if sys.platform == 'win32':
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(console_fmt)
    root.addHandler(ch)

    # Файл всех логов (DEBUG+)
    fh_all = logging.FileHandler(
        LOG_DIR / f"macro_realtime_{datetime.now():%Y%m%d}.log",
        encoding='utf-8'
    )
    fh_all.setLevel(logging.DEBUG)
    fh_all.setFormatter(detailed_fmt)
    root.addHandler(fh_all)

    # Файл ошибок (WARNING+)
    fh_err = logging.FileHandler(
        LOG_DIR / f"macro_errors_{datetime.now():%Y%m%d}.log",
        encoding='utf-8'
    )
    fh_err.setLevel(logging.WARNING)
    fh_err.setFormatter(detailed_fmt)
    root.addHandler(fh_err)

    return logging.getLogger(__name__)


log = setup_logging()


# ======================================================================
#                         БАЗА ДАННЫХ
# ======================================================================

def get_engine():
    if not DB_URL:
        raise ValueError("DB_URL не установлен в .env")
    return create_engine(DB_URL, connect_args={"ssl_context": False})


def ensure_table(engine):
    """Создать таблицы macro (metadata) и macro_data (данные)."""
    with engine.connect() as conn:
        # Metadata-таблица (реестр макроиндикаторов)
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS macro (
                id SERIAL PRIMARY KEY,
                indicator VARCHAR(50) UNIQUE NOT NULL,
                name VARCHAR(255) NOT NULL,
                frequency VARCHAR(20) NOT NULL,
                source VARCHAR(100),
                start_date DATE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))

        # Таблица данных
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS macro_data (
                id SERIAL PRIMARY KEY,
                indicator VARCHAR(50) NOT NULL,
                period_date DATE NOT NULL,
                value NUMERIC(24, 2) NOT NULL,
                source VARCHAR(100),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(indicator, period_date)
            )
        """))
        conn.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_macro_data_indicator_date
            ON macro_data(indicator, period_date)
        """))

        # Заполнить macro реестр (upsert)
        conn.execute(text("""
            INSERT INTO macro (indicator, name, frequency, source)
            VALUES ('M2_MONTHLY', 'Денежная масса M2', 'monthly', 'CBR_XLSX')
            ON CONFLICT (indicator) DO UPDATE SET
                name = EXCLUDED.name,
                frequency = EXCLUDED.frequency,
                source = EXCLUDED.source
        """))
        conn.execute(text("""
            INSERT INTO macro (indicator, name, frequency, source)
            VALUES ('GDP_QUARTERLY', 'ВВП России', 'quarterly', 'ROSSTAT_XLSX')
            ON CONFLICT (indicator) DO UPDATE SET
                name = EXCLUDED.name,
                frequency = EXCLUDED.frequency,
                source = EXCLUDED.source
        """))

        conn.commit()


# ======================================================================
#                         ВАЛИДАЦИЯ ДАННЫХ
# ======================================================================

def validate_value(value: float, indicator: str) -> Tuple[bool, str]:
    """Проверка одного значения на допустимость."""
    if value is None:
        return False, "значение None"
    if value <= 0:
        return False, f"отрицательное или нулевое значение: {value}"

    if indicator == 'M2_MONTHLY':
        if value < M2_MIN_VALUE:
            return False, f"M2 слишком мало: {value} < {M2_MIN_VALUE}"
        if value > M2_MAX_VALUE:
            return False, f"M2 слишком велико: {value} > {M2_MAX_VALUE}"
    elif indicator == 'GDP_QUARTERLY':
        if value < GDP_MIN_VALUE:
            return False, f"GDP слишком мало: {value} < {GDP_MIN_VALUE}"
        if value > GDP_MAX_VALUE:
            return False, f"GDP слишком велико: {value} > {GDP_MAX_VALUE}"

    return True, "OK"


def validate_date(period_date: date, indicator: str) -> Tuple[bool, str]:
    """Проверка даты на допустимость."""
    today = date.today()

    if period_date > today + timedelta(days=90):
        return False, f"дата в далеком будущем: {period_date}"

    if indicator == 'M2_MONTHLY' and period_date < M2_MIN_DATE:
        return False, f"M2 до {M2_MIN_DATE}: {period_date}"
    elif indicator == 'GDP_QUARTERLY' and period_date < GDP_MIN_DATE:
        return False, f"GDP до {GDP_MIN_DATE}: {period_date}"

    return True, "OK"


def validate_series_jumps(data: List[Tuple[date, float]], indicator: str) -> List[str]:
    """
    Проверить серию на аномальные скачки между соседними значениями.
    Возвращает список предупреждений.
    """
    warnings = []
    max_jump = M2_MAX_JUMP_PCT if indicator == 'M2_MONTHLY' else GDP_MAX_JUMP_PCT
    sorted_data = sorted(data, key=lambda x: x[0])

    for i in range(1, len(sorted_data)):
        prev_date, prev_val = sorted_data[i - 1]
        curr_date, curr_val = sorted_data[i]

        if prev_val == 0:
            continue

        change_pct = abs(curr_val - prev_val) / prev_val * 100
        if change_pct > max_jump:
            warnings.append(
                f"Скачок {change_pct:.1f}% между {prev_date} ({prev_val:,.0f}) "
                f"и {curr_date} ({curr_val:,.0f})"
            )

    return warnings


# ======================================================================
#                    ПРОВЕРКА ДЫР В ДАННЫХ
# ======================================================================

def check_m2_gaps(engine) -> Dict:
    """Проверить M2 на пропущенные месяцы."""
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT period_date, value FROM macro_data
            WHERE indicator = 'M2_MONTHLY'
            ORDER BY period_date
        """)).fetchall()

    if not rows:
        return {'status': 'empty', 'total': 0, 'gaps': [], 'first': None, 'last': None}

    dates = [r[0] for r in rows]
    values = [float(r[1]) for r in rows]
    gaps = []

    for i in range(1, len(dates)):
        prev = dates[i - 1]
        curr = dates[i]

        # Ожидаемая следующая дата: конец следующего месяца
        if prev.month == 12:
            expected_year = prev.year + 1
            expected_month = 1
        else:
            expected_year = prev.year
            expected_month = prev.month + 1
        _, last_day = monthrange(expected_year, expected_month)
        expected = date(expected_year, expected_month, last_day)

        if curr != expected:
            # Посчитаем сколько месяцев пропущено
            months_diff = (curr.year - prev.year) * 12 + (curr.month - prev.month)
            if months_diff > 1:
                gaps.append({
                    'from': prev,
                    'to': curr,
                    'missing_months': months_diff - 1
                })

    # Проверка на аномальные скачки
    jump_warnings = validate_series_jumps(list(zip(dates, values)), 'M2_MONTHLY')

    return {
        'status': 'ok' if not gaps and not jump_warnings else 'issues',
        'total': len(dates),
        'first': dates[0],
        'last': dates[-1],
        'gaps': gaps,
        'jump_warnings': jump_warnings,
    }


def check_gdp_gaps(engine) -> Dict:
    """Проверить GDP на пропущенные кварталы."""
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT period_date, value FROM macro_data
            WHERE indicator = 'GDP_QUARTERLY'
            ORDER BY period_date
        """)).fetchall()

    if not rows:
        return {'status': 'empty', 'total': 0, 'gaps': [], 'first': None, 'last': None}

    dates = [r[0] for r in rows]
    values = [float(r[1]) for r in rows]
    gaps = []

    # Квартальные даты: 31.03, 30.06, 30.09, 31.12
    Q_SEQUENCE = [(3, 31), (6, 30), (9, 30), (12, 31)]

    for i in range(1, len(dates)):
        prev = dates[i - 1]
        curr = dates[i]

        # Ожидаемый следующий квартал
        prev_q_idx = None
        for idx, (m, _) in enumerate(Q_SEQUENCE):
            if prev.month == m:
                prev_q_idx = idx
                break

        if prev_q_idx is not None:
            next_q_idx = (prev_q_idx + 1) % 4
            next_year = prev.year + (1 if next_q_idx == 0 else 0)
            next_m, next_d = Q_SEQUENCE[next_q_idx]
            expected = date(next_year, next_m, next_d)

            if curr != expected:
                quarters_diff = (curr.year - prev.year) * 4
                for idx, (m, _) in enumerate(Q_SEQUENCE):
                    if curr.month == m:
                        quarters_diff += idx
                        break
                quarters_diff -= prev_q_idx
                if quarters_diff > 1:
                    gaps.append({
                        'from': prev,
                        'to': curr,
                        'missing_quarters': quarters_diff - 1
                    })

    # Проверка на аномальные скачки
    jump_warnings = validate_series_jumps(list(zip(dates, values)), 'GDP_QUARTERLY')

    return {
        'status': 'ok' if not gaps and not jump_warnings else 'issues',
        'total': len(dates),
        'first': dates[0],
        'last': dates[-1],
        'gaps': gaps,
        'jump_warnings': jump_warnings,
    }


def print_check_report(engine):
    """Вывести полный отчёт о целостности данных."""
    sep = '=' * 60
    log.info(sep)
    log.info("ПРОВЕРКА ЦЕЛОСТНОСТИ МАКРОДАННЫХ")
    log.info(sep)

    # --- M2 ---
    m2 = check_m2_gaps(engine)
    log.info("")
    log.info("--- M2_MONTHLY ---")
    if m2['status'] == 'empty':
        log.warning("  НЕТ ДАННЫХ M2 в БД!")
    else:
        log.info(f"  Записей: {m2['total']}")
        log.info(f"  Период: {m2['first']} - {m2['last']}")
        if m2['gaps']:
            log.warning(f"  ДЫРЫ: {len(m2['gaps'])} пропусков:")
            for g in m2['gaps']:
                log.warning(f"    {g['from']} -> {g['to']} ({g['missing_months']} мес. пропущено)")
        else:
            log.info("  Дыр нет (все месяцы подряд)")
        if m2.get('jump_warnings'):
            log.warning(f"  АНОМАЛИИ: {len(m2['jump_warnings'])} скачков:")
            for w in m2['jump_warnings']:
                log.warning(f"    {w}")
        else:
            log.info("  Аномальных скачков нет")

    # --- GDP ---
    gdp = check_gdp_gaps(engine)
    log.info("")
    log.info("--- GDP_QUARTERLY ---")
    if gdp['status'] == 'empty':
        log.warning("  НЕТ ДАННЫХ GDP в БД!")
    else:
        log.info(f"  Записей: {gdp['total']}")
        log.info(f"  Период: {gdp['first']} - {gdp['last']}")
        if gdp['gaps']:
            log.warning(f"  ДЫРЫ: {len(gdp['gaps'])} пропусков:")
            for g in gdp['gaps']:
                log.warning(f"    {g['from']} -> {g['to']} ({g['missing_quarters']} кварт. пропущено)")
        else:
            log.info("  Дыр нет (все кварталы подряд)")
        if gdp.get('jump_warnings'):
            log.warning(f"  АНОМАЛИИ: {len(gdp['jump_warnings'])} скачков:")
            for w in gdp['jump_warnings']:
                log.warning(f"    {w}")
        else:
            log.info("  Аномальных скачков нет")

    # --- Актуальность ---
    log.info("")
    log.info("--- АКТУАЛЬНОСТЬ ---")
    today = date.today()
    if m2['last']:
        m2_age = (today - m2['last']).days
        if m2_age > 60:
            log.warning(f"  M2: последняя запись {m2['last']} ({m2_age} дней назад) — УСТАРЕЛА")
        else:
            log.info(f"  M2: последняя запись {m2['last']} ({m2_age} дней назад) — OK")
    if gdp['last']:
        gdp_age = (today - gdp['last']).days
        if gdp_age > 120:
            log.warning(f"  GDP: последняя запись {gdp['last']} ({gdp_age} дней назад) — УСТАРЕЛА")
        else:
            log.info(f"  GDP: последняя запись {gdp['last']} ({gdp_age} дней назад) — OK")

    log.info(sep)

    return m2, gdp


# ═══════════════════════════════════════════════════════════════════
# M2: Автозагрузка с ЦБ РФ
# ═══════════════════════════════════════════════════════════════════

def fetch_m2_from_cbr(engine) -> int:
    """Скачать M2 из XLSX ЦБ РФ и обновить в БД."""
    log.info("Загрузка M2 с ЦБ РФ...")

    # Скачиваем XLSX во временный файл
    tmp_path = None
    try:
        log.debug(f"  URL: {CBR_M2_URL}")
        req = urllib.request.Request(CBR_M2_URL, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
        })
        resp = urllib.request.urlopen(req, timeout=30)
        status = resp.getcode()
        log.info(f"  HTTP статус: {status}")

        if status != 200:
            log.error(f"  Неожиданный HTTP статус: {status}")
            return 0

        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        raw_data = resp.read()
        file_size = len(raw_data)
        with os.fdopen(tmp_fd, 'wb') as f:
            f.write(raw_data)

        log.info(f"  XLSX скачан: {file_size:,} байт")

        if file_size < 1000:
            log.error(f"  Файл слишком маленький ({file_size} байт) — вероятно ошибка загрузки")
            return 0

        # Парсим
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)

        sheet_name = 'Денежные агрегаты'
        if sheet_name not in wb.sheetnames:
            log.error(f"  Лист '{sheet_name}' не найден! Доступные: {wb.sheetnames}")
            wb.close()
            return 0

        ws = wb[sheet_name]

        # Row 1 = даты, Row 14 = M2
        dates_row = []
        m2_row = []
        for cell in ws[1]:
            dates_row.append(cell.value)
        for cell in ws[14]:
            m2_row.append(cell.value)
        wb.close()

        log.debug(f"  Столбцов с датами: {len(dates_row)}, с M2: {len(m2_row)}")

        # Парсим пары (date, value)
        data = []
        skipped = 0
        validation_errors = []

        for dt, val in zip(dates_row[1:], m2_row[1:]):  # skip col A (label)
            if dt is None or val is None:
                skipped += 1
                continue
            try:
                if isinstance(dt, datetime):
                    d = dt.date()
                elif isinstance(dt, date):
                    d = dt
                else:
                    skipped += 1
                    continue

                _, last_day = monthrange(d.year, d.month)
                period_date = date(d.year, d.month, last_day)
                value = float(val)

                # Валидация даты
                ok, msg = validate_date(period_date, 'M2_MONTHLY')
                if not ok:
                    validation_errors.append(f"  Дата: {msg}")
                    continue

                # Валидация значения
                ok, msg = validate_value(value, 'M2_MONTHLY')
                if not ok:
                    validation_errors.append(f"  Значение {period_date}: {msg}")
                    continue

                data.append((period_date, value))
            except (ValueError, TypeError) as e:
                validation_errors.append(f"  Ошибка парсинга: {dt} / {val}: {e}")
                continue

        if not data:
            log.error("  Не удалось распарсить ни одной записи M2!")
            for err in validation_errors[:10]:
                log.error(err)
            return 0

        log.info(f"  Распарсено: {len(data)} записей M2 ({data[0][0]} - {data[-1][0]})")
        if skipped:
            log.debug(f"  Пропущено пустых: {skipped}")
        if validation_errors:
            log.warning(f"  Ошибок валидации: {len(validation_errors)}")
            for err in validation_errors[:5]:
                log.warning(err)

        # Проверка на аномальные скачки
        jumps = validate_series_jumps(data, 'M2_MONTHLY')
        if jumps:
            log.warning(f"  Обнаружены аномальные скачки в M2 ({len(jumps)}):")
            for j in jumps[:5]:
                log.warning(f"    {j}")

        # Получаем последнюю дату в БД
        with engine.connect() as conn:
            r = conn.execute(text(
                "SELECT MAX(period_date) FROM macro_data WHERE indicator = 'M2_MONTHLY'"
            )).fetchone()
            last_db_date = r[0] if r and r[0] else None

            r2 = conn.execute(text(
                "SELECT COUNT(*) FROM macro_data WHERE indicator = 'M2_MONTHLY'"
            )).fetchone()
            db_count_before = r2[0] if r2 else 0

        log.info(f"  В БД до обновления: {db_count_before} записей, последняя: {last_db_date}")

        # Вставляем/обновляем
        inserted = 0
        errors = 0
        with engine.connect() as conn:
            for period_date, value in data:
                try:
                    conn.execute(text("""
                        INSERT INTO macro_data (indicator, period_date, value, source)
                        VALUES ('M2_MONTHLY', :pd, :val, 'CBR_XLSX')
                        ON CONFLICT (indicator, period_date) DO UPDATE SET
                            value = EXCLUDED.value, source = EXCLUDED.source,
                            created_at = CURRENT_TIMESTAMP
                    """), {"pd": period_date, "val": value})
                    inserted += 1
                except Exception as e:
                    errors += 1
                    log.error(f"  Ошибка вставки M2 {period_date}: {e}")
            conn.commit()

        # Верификация после вставки
        with engine.connect() as conn:
            r = conn.execute(text(
                "SELECT COUNT(*) FROM macro_data WHERE indicator = 'M2_MONTHLY'"
            )).fetchone()
            db_count_after = r[0] if r else 0

        new_count = len([d for d in data if last_db_date is None or d[0] > last_db_date])
        log.info(f"  M2 итог: {inserted} upserted, {new_count} новых, {errors} ошибок")
        log.info(f"  В БД после обновления: {db_count_after} записей (было {db_count_before})")

        if errors > 0:
            log.warning(f"  {errors} записей M2 не удалось вставить!")

        return inserted

    except urllib.error.URLError as e:
        log.error(f"  Ошибка сети при загрузке M2: {e}")
        return 0
    except Exception as e:
        log.error(f"  Ошибка загрузки M2: {e}", exc_info=True)
        return 0
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


# ═══════════════════════════════════════════════════════════════════
# GDP: Авто-загрузка с rosstat.gov.ru
# ═══════════════════════════════════════════════════════════════════

# Регулярка ловит ссылку вида "/storage/mediabank/VVP_kvartal_s_1995-2025.xlsx"
# (с опциональным подкаталогом-хешем, как у некоторых файлов на сайте Росстата).
# Разделитель после "s" плавает: было "s_1995" и "s1995", с 2026 Росстат стал
# писать "VVP_kvartal_s-1995-2026.xlsx" (ДЕФИС) → из-за этого discover отвалился
# 02.07 и ВВП завис на Q4 2025. Допускаем `-`/`_`/пусто: `s[-_]?\d{4}`.
GDP_XLSX_HREF_RE = _re.compile(
    r'href="(/storage/mediabank/(?:[^"/]+/)?VVP_kvartal_s[-_]?\d{4}[^"]*\.xlsx)"',
    _re.IGNORECASE,
)
# "Обновлено 10.04.2026г." на странице "Содержание" в самом xlsx.
GDP_PUBLISH_DATE_RE = _re.compile(r'Обновлено\s+(\d{1,2})\.(\d{1,2})\.(\d{4})')


def discover_gdp_url() -> Optional[str]:
    """Парсит landing-страницу Росстата и возвращает абсолютный URL квартального
    файла VVP_kvartal_s_*.xlsx. None если ссылка не найдена.

    Использует regex по HTML, а не BeautifulSoup — экономим зависимости.
    """
    log.info("Discover URL квартального ВВП на rosstat.gov.ru...")
    try:
        req = urllib.request.Request(
            ROSSTAT_GDP_PAGE_URL,
            headers={"User-Agent": "Mozilla/5.0 (X11; Linux x86_64)"},
        )
        with urllib.request.urlopen(req, timeout=30, context=_rosstat_ssl_context()) as resp:
            if resp.status != 200:
                log.error(f"  HTTP {resp.status} при загрузке {ROSSTAT_GDP_PAGE_URL}")
                return None
            html = resp.read().decode("utf-8", errors="replace")
    except Exception as e:
        log.error(f"  Не удалось загрузить landing-страницу: {e}{_tls_hint(e)}")
        return None

    m = GDP_XLSX_HREF_RE.search(html)
    if not m:
        log.error("  Ссылка на VVP_kvartal_s_*.xlsx не найдена на странице. "
                  "Возможно, Росстат поменял layout — проверьте вручную.")
        return None

    rel_path = m.group(1)
    abs_url = ROSSTAT_BASE_URL + rel_path
    log.info(f"  Найдена ссылка: {abs_url}")
    return abs_url


def _download_to_tempfile(url: str, ssl_context=None) -> Optional[str]:
    """Скачивает URL во временный .xlsx, возвращает путь или None."""
    try:
        req = urllib.request.Request(
            url, headers={"User-Agent": "Mozilla/5.0 (X11; Linux x86_64)"}
        )
        with urllib.request.urlopen(req, timeout=60, context=ssl_context) as resp:
            if resp.status != 200:
                log.error(f"  HTTP {resp.status} при загрузке {url}")
                return None
            data = resp.read()
        if len(data) < 5_000:
            log.error(f"  Файл слишком маленький ({len(data)} байт)")
            return None

        fd, path = tempfile.mkstemp(suffix=".xlsx")
        with os.fdopen(fd, "wb") as f:
            f.write(data)
        log.info(f"  Скачано: {len(data):,} байт → {path}")
        return path
    except Exception as e:
        log.error(f"  Ошибка скачивания {url}: {e}{_tls_hint(e)}")
        return None


def parse_gdp_xlsx(xlsx_path: str) -> Tuple[Dict[date, float], List[str], Optional[date]]:
    """Парсит Excel Росстата с квартальным ВВП.

    Берёт ТОЛЬКО листы, у которых R2 содержит «в текущих ценах» —
    игнорирует индексы физ. объёма, дефляторы, постоянные цены.

    Возвращает (data, validation_errors, publish_date).
      - data: {period_date: value_млрд_руб}
      - validation_errors: список строк с диагностикой
      - publish_date: дата «Обновлено DD.MM.YYYY» с листа Содержание (или None)
    """
    import openpyxl

    Q_DATES = {
        'I квартал': (3, 31), 'II квартал': (6, 30),
        'III квартал': (9, 30), 'IV квартал': (12, 31),
    }

    if not os.path.exists(xlsx_path):
        return {}, [f"Файл не найден: {xlsx_path}"], None

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        return {}, [f"Не удалось открыть xlsx: {e}"], None

    all_data: Dict[date, float] = {}
    validation_errors: List[str] = []
    sheets_used: List[str] = []
    publish_date: Optional[date] = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_list = list(ws.iter_rows(min_row=1, max_row=30, values_only=True))
        if len(rows_list) < 5:
            continue

        # Лист "Содержание" — найдём дату публикации "Обновлено DD.MM.YYYY"
        if sheet_name.lower().startswith('содержание') or 'содержание' in str(rows_list[0][0] or '').lower():
            for row in rows_list:
                for cell in row:
                    if not cell:
                        continue
                    m = GDP_PUBLISH_DATE_RE.search(str(cell))
                    if m:
                        try:
                            publish_date = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
                        except ValueError:
                            pass
                        break
            continue

        # Фильтр листа: R2 должен говорить «в текущих ценах»
        title = str(rows_list[1][0] or '') if rows_list[1] else ''
        if 'в текущих ценах' not in title:
            continue
        sheets_used.append(sheet_name)

        # Парсинг как раньше: R3=год, R4=квартал, R5=значение
        row3 = rows_list[2]
        row4 = rows_list[3]
        row5 = rows_list[4]

        for col_idx in range(len(row4)):
            quarter_label = row4[col_idx]
            value = row5[col_idx]
            if quarter_label not in Q_DATES or value is None:
                continue

            # Год — первая непустая ячейка в R3 при сканировании влево
            year = None
            for j in range(col_idx, -1, -1):
                y = row3[j]
                if y is not None:
                    try:
                        year = int(str(y).strip()[:4])
                        break
                    except (ValueError, TypeError):
                        continue
            if year is None:
                continue

            month, day = Q_DATES[quarter_label]
            period_date = date(year, month, day)
            try:
                fval = float(value)
            except (ValueError, TypeError):
                continue

            ok, msg = validate_date(period_date, 'GDP_QUARTERLY')
            if not ok:
                validation_errors.append(f"Дата: {msg}")
                continue
            ok, msg = validate_value(fval, 'GDP_QUARTERLY')
            if not ok:
                validation_errors.append(f"Значение {period_date}: {msg}")
                continue

            # Последний лист wins на стыке (например, лист 2 'переписывает' лист 1
            # для года 2011, что соответствует более актуальным оценкам Росстата).
            all_data[period_date] = fval

    wb.close()

    if not sheets_used:
        validation_errors.append(
            "Не найдено ни одного листа с «в текущих ценах» — изменилась структура файла Росстата?"
        )
    return all_data, validation_errors, publish_date


def _upsert_gdp_data(engine, all_data: Dict[date, float], source_tag: str) -> Tuple[int, int]:
    """Записать GDP-серию в БД (upsert, без DELETE — ревизии затираются точечно).

    Возвращает (inserted_or_updated, errors).
    """
    inserted, errors = 0, 0
    with engine.connect() as conn:
        for period_date, value in sorted(all_data.items()):
            try:
                conn.execute(text("""
                    INSERT INTO macro_data (indicator, period_date, value, source)
                    VALUES ('GDP_QUARTERLY', :pd, :val, :src)
                    ON CONFLICT (indicator, period_date) DO UPDATE SET
                        value = EXCLUDED.value,
                        source = EXCLUDED.source,
                        created_at = CURRENT_TIMESTAMP
                """), {"pd": period_date, "val": value, "src": source_tag})
                inserted += 1
            except Exception as e:
                errors += 1
                log.error(f"  Ошибка вставки GDP {period_date}: {e}")
        conn.commit()
    return inserted, errors


def _diff_with_db(engine, new_data: Dict[date, float]) -> Tuple[List[date], List[Tuple[date, float, float]]]:
    """Сравнивает новые данные с тем что уже в БД.

    Возвращает (новые_кварталы, ревизированные_кварталы).
    Ревизированный = существует в БД но значение отличается > 0.5%.
    """
    with engine.connect() as conn:
        rows = conn.execute(text(
            "SELECT period_date, value FROM macro_data WHERE indicator = 'GDP_QUARTERLY'"
        )).fetchall()
    db_map = {r[0]: float(r[1]) for r in rows}

    new_dates: List[date] = []
    revised: List[Tuple[date, float, float]] = []
    for d, v in sorted(new_data.items()):
        if d not in db_map:
            new_dates.append(d)
        else:
            old = db_map[d]
            if old == 0:
                continue
            if abs(v - old) / abs(old) > 0.005:  # >0.5%
                revised.append((d, old, v))
    return new_dates, revised


def fetch_gdp_from_rosstat(engine) -> int:
    """Полный auto-pipeline: discover URL → download → parse → diff → upsert.

    Возвращает количество кварталов в полученной серии. 0 = ошибка/нет данных.
    """
    log.info("Загрузка GDP с rosstat.gov.ru...")

    url = discover_gdp_url()
    if not url:
        log.warning("  Авто-загрузка GDP не удалась — используйте --import-gdp FILE.xlsx")
        return 0

    tmp_path = _download_to_tempfile(url, ssl_context=_rosstat_ssl_context())
    if not tmp_path:
        log.warning("  Скачивание не удалось — используйте --import-gdp FILE.xlsx")
        return 0

    try:
        all_data, errors, publish_date = parse_gdp_xlsx(tmp_path)
        if not all_data:
            log.error("  Парсинг не дал ни одного квартала.")
            for err in errors[:5]:
                log.error(f"    {err}")
            return 0

        if publish_date:
            log.info(f"  Дата публикации Росстата: {publish_date}")
        log.info(f"  Распарсено: {len(all_data)} кварталов "
                 f"({min(all_data)} → {max(all_data)})")
        if errors:
            log.warning(f"  Ошибок валидации: {len(errors)} (первые 3 ниже)")
            for err in errors[:3]:
                log.warning(f"    {err}")

        # Diff с БД — что нового, что переписали
        new_dates, revised = _diff_with_db(engine, all_data)
        if new_dates:
            log.info(f"  НОВЫЕ кварталы: {len(new_dates)}")
            for d in new_dates[-4:]:  # покажем последние 4
                log.info(f"    + {d}: {all_data[d]:>12,.2f} млрд ₽")
        if revised:
            log.info(f"  РЕВИЗИИ Росстата (>0.5%): {len(revised)}")
            for d, old, new in revised[-4:]:
                pct = (new - old) / old * 100
                log.info(f"    ~ {d}: {old:>12,.2f} → {new:>12,.2f} ({pct:+.2f}%)")
        if not new_dates and not revised:
            log.info("  Изменений нет — БД уже актуальна.")

        # Аномальные скачки
        jumps = validate_series_jumps(sorted(all_data.items()), 'GDP_QUARTERLY')
        if jumps:
            log.warning(f"  Аномальные скачки: {len(jumps)}")
            for j in jumps[-3:]:
                log.warning(f"    {j}")

        # Upsert
        inserted, db_errors = _upsert_gdp_data(engine, all_data, source_tag='ROSSTAT_AUTO')
        log.info(f"  GDP upsert: {inserted}/{len(all_data)}, ошибок: {db_errors}")
        return inserted

    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


# ═══════════════════════════════════════════════════════════════════
# GDP: Импорт из Excel (ручной — fallback)
# ═══════════════════════════════════════════════════════════════════

def import_gdp_from_xlsx(engine, xlsx_path: str) -> int:
    """Импорт GDP из локального Excel (fallback если auto-fetch не работает)."""
    if not os.path.exists(xlsx_path):
        log.error(f"Файл не найден: {xlsx_path}")
        return 0

    file_size = os.path.getsize(xlsx_path)
    log.info(f"Импорт GDP из {xlsx_path} ({file_size:,} байт)...")

    all_data, errors, publish_date = parse_gdp_xlsx(xlsx_path)
    if publish_date:
        log.info(f"  Дата публикации Росстата: {publish_date}")

    if not all_data:
        log.error("  Не удалось распарсить ни одной записи GDP!")
        for err in errors[:10]:
            log.error(f"    {err}")
        return 0

    log.info(f"  Распарсено: {len(all_data)} кварталов ({min(all_data)} → {max(all_data)})")
    if errors:
        log.warning(f"  Ошибок валидации: {len(errors)}")
        for err in errors[:5]:
            log.warning(f"    {err}")

    # Аномальные скачки
    jumps = validate_series_jumps(sorted(all_data.items()), 'GDP_QUARTERLY')
    if jumps:
        log.warning(f"  Аномальные скачки в GDP: {len(jumps)}")
        for j in jumps[:5]:
            log.warning(f"    {j}")

    # Diff с БД
    new_dates, revised = _diff_with_db(engine, all_data)
    if new_dates:
        log.info(f"  НОВЫЕ кварталы: {len(new_dates)} (последний: {new_dates[-1]})")
    if revised:
        log.info(f"  Ревизий Росстата: {len(revised)}")

    inserted, db_errors = _upsert_gdp_data(engine, all_data, source_tag='ROSSTAT_XLSX')
    log.info(f"  GDP итог: {inserted}/{len(all_data)} upsert, {db_errors} ошибок")
    return inserted


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

async def update_macro(force: bool = False) -> dict:
    """Обновить все макроданные."""
    engine = get_engine()
    ensure_table(engine)
    results = {}

    # M2 — всегда обновляем (ЦБ обновляет ежемесячно)
    m2_count = fetch_m2_from_cbr(engine)
    results['M2'] = m2_count

    # GDP — теперь auto. Росстат публикует Q+0 квартал через ~2 месяца после
    # окончания квартала. Скрипт сам найдёт URL на /statistics/accounts.
    # Failure здесь не критичен — есть fallback `--import-gdp FILE`.
    try:
        gdp_count = fetch_gdp_from_rosstat(engine)
        results['GDP'] = gdp_count
    except Exception as e:
        log.error(f"GDP auto-fetch упал: {e}", exc_info=True)
        results['GDP'] = 0

    # KEY_RATE — дневной ряд ставки ЦБ (Candles/fetch_key_rate.py). До 2026-07-21
    # запускался только вручную и молча застревал (ряд простоял 40 дней и проспал
    # снижение 14.50→14.25). Окно 30 дней, upsert идемпотентен. Failure не
    # критичен: свежесть ловит /api/health/data (kind=t_plus_1).
    try:
        proc = subprocess.run(
            [sys.executable, str(PROJECT_DIR / 'Candles' / 'fetch_key_rate.py'),
             '--from', (date.today() - timedelta(days=30)).isoformat()],
            capture_output=True, text=True, timeout=120,
        )
        tail = (proc.stdout or proc.stderr or '').strip().splitlines()
        log.info(f"KEY_RATE: rc={proc.returncode} — {tail[-1] if tail else 'нет вывода'}")
        results['KEY_RATE'] = 1 if proc.returncode == 0 else 0
    except Exception as e:
        log.error(f"KEY_RATE fetch упал: {e}")
        results['KEY_RATE'] = 0

    # Проверка целостности после обновления
    log.info("Проверка целостности данных после обновления...")
    m2_check = check_m2_gaps(engine)
    gdp_check = check_gdp_gaps(engine)

    if m2_check['status'] == 'empty':
        log.warning("  M2: данных нет!")
    elif m2_check['gaps']:
        log.warning(f"  M2: обнаружено {len(m2_check['gaps'])} дыр в данных")
    else:
        log.info(f"  M2: OK ({m2_check['total']} записей, {m2_check['first']} - {m2_check['last']})")

    if gdp_check['status'] == 'empty':
        log.warning("  GDP: данных нет (требуется ручной импорт --import-gdp)")
    elif gdp_check['gaps']:
        log.warning(f"  GDP: обнаружено {len(gdp_check['gaps'])} дыр в данных")
    else:
        log.info(f"  GDP: OK ({gdp_check['total']} записей, {gdp_check['first']} - {gdp_check['last']})")

    results['m2_check'] = m2_check['status']
    results['gdp_check'] = gdp_check['status']

    engine.dispose()
    return results


async def main():
    parser = argparse.ArgumentParser(description="Macro data realtime updater")
    parser.add_argument("--once", action="store_true", help="Single update")
    parser.add_argument("--force", action="store_true", help="Force full reload")
    parser.add_argument("--import-gdp", type=str, metavar="FILE", help="Import GDP from Excel")
    parser.add_argument("--check", action="store_true", help="Check data integrity only")
    args = parser.parse_args()

    sep = '=' * 60
    log.info(sep)
    log.info("MACRO DATA UPDATER")
    log.info(f"Time: {datetime.now()}")
    log.info(f"Args: once={args.once}, force={args.force}, check={args.check}")
    log.info(sep)

    engine = get_engine()
    ensure_table(engine)

    # Режим проверки целостности
    if args.check:
        print_check_report(engine)
        engine.dispose()
        return

    # Импорт GDP
    if args.import_gdp:
        count = import_gdp_from_xlsx(engine, args.import_gdp)
        if count > 0:
            log.info("Запуск проверки целостности после импорта...")
            print_check_report(engine)
        engine.dispose()
        return

    # Однократное обновление
    if args.once:
        is_trading, reason = is_trading_day()
        if not is_trading and not args.force:
            log.info(f"Skip: {reason} (use --force)")
            print(json.dumps({"пропуск": reason}, ensure_ascii=False))
            return
        results = await update_macro(force=args.force)
        for k, v in results.items():
            log.info(f"  {k}: {v}")
    else:
        # Daemon mode
        log.info("Daemon mode: checking daily at 09:30 MSK")
        last_update = None
        while True:
            now = get_moscow_time()
            today = now.date()
            is_trade, _ = is_trading_day(today)

            if is_trade and now.hour >= 9 and last_update != today:
                log.info(f"Daily update ({now:%H:%M})...")
                try:
                    await update_macro()
                except Exception as e:
                    log.error(f"Ошибка daily update: {e}", exc_info=True)
                last_update = today

            await asyncio.sleep(300)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        log.info("Stopped")
    except Exception as e:
        log.critical(f"Error: {e}", exc_info=True)
        sys.exit(1)
